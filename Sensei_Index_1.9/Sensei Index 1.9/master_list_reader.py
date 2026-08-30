# -*- coding: utf-8 -*-
"""
Parser for Instrumentation_Master_List.xlsx (Sensei Index 2.1, Phase 10).

Standalone module - only depends on openpyxl and re, not on data_access or
any GUI code - so it's importable and testable with a hand-built workbook
and nothing else running. data_access.py is the orchestration layer on top
(load_master_list / save_master_list / master_list_items / reconcile_
master_list); this module only knows how to turn one .xlsx into normalized
item dicts.

Every defensive choice below traces to a verified finding from directly
inspecting the real Instrumentation_Master_List.xlsx (see the v2.1 spec,
section 1) - cited inline so the "why" survives:

1. Sheets and true row counts. `KIB AREA 29203 FLOWLINE`'s dimension record
   claims 1,048,575 rows for ~20 real ones - iterating ws.max_row (or, for
   that matter, trusting ANY sheet's claimed dimension) would hang or waste
   a full sheet scan. `_iter_data_rows` below stops after BLANK_STREAK_STOP
   consecutive fully-blank rows instead. (`9050 SUS` has the same problem
   on a smaller scale: its used-range claims 1,383 rows for 134 real ones -
   confirmed by direct inspection of the uploaded workbook. The spec's own
   "~1,380 rows" acceptance-criteria estimate for that sheet conflates the
   inflated claimed dimension with the real row count; the real count is
   134, and that's what this parser - correctly - produces.)
2. Headers are inconsistent across sheets: four sheets use "Tag Number"
   (one with a trailing space, `KIB AREA 29203 FLOWLINE`); `9050 SUS` uses
   `Tag_Number`, `Instrument_Type_Desc`, `Model_Name`, `Manufacturer_Name`,
   and a P&ID header containing non-breaking spaces. `Remarks` vs `Remark`
   also varies. `normalize_header` below never exact-matches raw header
   text for this reason.
3. Tag numbers contain internal padding: `29103-AT  -2720A`,
   `29103-PIT -2171` - the tag-type segment is space-padded to a fixed
   width, while the tracker's own tags are unpadded (`29103-FIT-0916`).
   `canonical_tag` is the one function used for every cross-file tag
   comparison in this app.
4. Excel float-coercion on model numbers: WIKA gauges show Model `233.34`,
   Leser PSVs show long decimal models. `format_model` renders a float
   without introducing binary-float noise and always returns a string.
5. Typos exist in type descriptions (`MOSISTURE`, `TEMPTERATURE`,
   `UPSTREAN`) - never "fixed" (it's the client's data), which is exactly
   why `classify_kind` keys on substrings like "TRANSMITTER"/"TX" rather
   than exact strings that a typo would silently miss.
6. Received / Installed columns are sparse "Yes"/blank, same blank-means-
   unset convention as equipment_status.json. `parse_yes` maps "Yes"
   (case-insensitive, stripped) to True, anything else to False.
7. Most master-list rows are not transmitters or valves - gauges,
   switches, thermowells, PSVs, restriction orifices, flow elements,
   positioners, regulators, motor trips all appear. `classify_kind` sorts
   every row into transmitter / valve / out_of_scope so Phase 11's
   coverage math never penalizes the user for a "missing" thermowell.
8. Master-list sheets map onto areas/series the tracker already knows, but
   not always by an exact number match (`9050 SUS` vs registered series
   `500`) - `guess_series_mapping` only auto-fills an exact match and
   leaves the rest for the user to confirm (data_access.py / the import
   dialog persist that choice in sheet_series_map).
"""
import re

import openpyxl

# ---------------------------------------------------------------------------
# Defensive scan limits
# ---------------------------------------------------------------------------
HEADER_SCAN_ROWS = 5     # header row is somewhere in the first 5 rows
BLANK_STREAK_STOP = 50   # stop a sheet after this many consecutive blank rows


# ---------------------------------------------------------------------------
# canonical_tag / parse_area_code
#
# These live here (not duplicated in data_access.py) but ARE re-exported
# from data_access.py (`from master_list_reader import canonical_tag,
# parse_area_code`) - that's the public home the v2.1 spec's Phase 10.2
# describes, and where v2.0 Phase 6 is meant to find/reuse parse_area_code.
# They live in this module, not data_access.py, purely to avoid a circular
# import (data_access.py already needs to import this module for the
# parser itself).
# ---------------------------------------------------------------------------
def canonical_tag(value):
    """One canonicalization used for EVERY cross-file tag comparison in
    this app: strip, remove whitespace adjacent to hyphens (defuses the
    master list's tag-type padding, e.g. '29103-AT  -2720A'), collapse any
    other stray internal whitespace, uppercase. The original string should
    always be kept alongside this for display - this is a comparison key,
    not a display value."""
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"\s+", "", text)
    return text.upper()


def parse_area_code(tag):
    """The prefix before the first '-' in a (canonical or raw) tag, e.g.
    '29103-PIT-2171' -> '29103'. '' for a tag with no hyphen at all."""
    if not tag:
        return ""
    return str(tag).split("-", 1)[0]


def tag_type_of(canonical):
    """The segment between the first and second hyphen of a canonical tag,
    e.g. '29103-PIT-2171' -> 'PIT'. '' if there's no second segment."""
    parts = str(canonical or "").split("-")
    return parts[1] if len(parts) >= 2 else ""


# ---------------------------------------------------------------------------
# Header normalization + the alias table (finding 2)
# ---------------------------------------------------------------------------
def normalize_header(value):
    """casefold, strip, collapse runs of whitespace/underscore/non-breaking
    space to a single space. 'Tag_Number' / 'Tag Number' / 'Tag Number ' /
    'P\\xa0\\xa0\\xa0\\xa0\\xa0ID' all normalize to something this module's
    alias table recognizes - never exact-match raw header text."""
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"[\s_]+", " ", text)
    return text.strip().casefold()


HEADER_ALIASES = {
    "tag number": "tag",
    "service": "service",
    "instrument type description": "type_desc",
    "instrument type desc": "type_desc",
    "p&id": "pid",
    "p id": "pid",
    "pid": "pid",
    "line number": "line_number",
    "location layout": "location_layout",
    "module id": "module_id",
    "issued for": "issued_for",
    "received": "received",
    "installed": "installed",
    "ewp": "ewp",
    "model": "model",
    "model name": "model",
    "manufacturer": "manufacturer",
    "manufacturer name": "manufacturer",
    "installation details": "install_details",
    "remarks": "remarks",
    "remark": "remarks",
}

# Minimum a sheet must have to be worth importing at all (10.2).
REQUIRED_KEYS = {"tag", "type_desc"}


# ---------------------------------------------------------------------------
# Model float-coercion repair (finding 4)
# ---------------------------------------------------------------------------
def format_model(value):
    """Excel stores some Model cells as numbers (WIKA '233.34', Leser
    '5263.5123...'). Returns a string always; for a float, uses Python's
    repr() (the shortest string that round-trips to the exact same float -
    no '...000000003' noise) rather than str()/format specs that can
    reintroduce it, and drops the decimal entirely when the value is a
    whole number."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else repr(value)
    return str(value).strip()


def parse_yes(value):
    """finding 6: blank-means-unset, 'Yes' (case/whitespace-insensitive)
    means set - same convention equipment_status.json already uses."""
    return str(value or "").strip().casefold() == "yes"


# ---------------------------------------------------------------------------
# Kind classification (10.3)
# ---------------------------------------------------------------------------
TRANSMITTER_TYPE_DESC_KEYWORDS = ["TRANSMITTER", "ANALYZER TX"]
TRANSMITTER_TAG_TYPES = {
    "PIT", "TIT", "LIT", "FIT", "AT", "PDIT", "TDT", "TT", "PT", "LT", "FT",
    "DPIT", "AIT", "ZIT",
}

VALVE_TYPE_DESC_KEYWORDS = ["CONTROL VALVE", "ON/OFF", "SHUTDOWN VALVE"]
VALVE_TAG_TYPES = {"XV", "LV", "PV", "FV", "KV", "TV", "HV", "PCV"}
# Relief devices - never valves, no matter what a tag-type fallback might
# suggest (10.3's explicit exclusion).
RELIEF_TAG_TYPES = {"PSV", "PSE"}
# Self-contained regulators read as valve-ish by tag (PCV/TCV) but are
# explicitly out_of_scope (10.3).
SELF_CONTAINED_KEYWORDS = [
    "REGULATOR SELF CONTAINED", "REGULATOR - SELF CONTAINED",
    "SELF CONTAINED REGULATOR", "SELF-CONTAINED",
]


def classify_kind(tag_type, type_desc):
    """"transmitter" | "valve" | "out_of_scope", by an ordered rule table:
    substring match on the uppercased type description first, tag-type
    fallback second (10.3). Never mutates/"fixes" a typo'd type_desc
    (finding 5) - substring matching on TRANSMITTER/TX tolerates
    'ANALYZER TX - MOSISTURE' the same as a correctly spelled one."""
    type_desc_up = (type_desc or "").upper()
    tag_type_up = (tag_type or "").upper()

    if any(kw in type_desc_up for kw in SELF_CONTAINED_KEYWORDS):
        return "out_of_scope"
    if tag_type_up in RELIEF_TAG_TYPES:
        return "out_of_scope"

    if any(kw in type_desc_up for kw in TRANSMITTER_TYPE_DESC_KEYWORDS):
        return "transmitter"
    if any(kw in type_desc_up for kw in VALVE_TYPE_DESC_KEYWORDS):
        return "valve"

    if tag_type_up in TRANSMITTER_TAG_TYPES:
        return "transmitter"
    if tag_type_up in VALVE_TAG_TYPES:
        return "valve"
    return "out_of_scope"


# ---------------------------------------------------------------------------
# Sheet parsing
# ---------------------------------------------------------------------------
def _is_blank_row(row):
    return all(c is None or (isinstance(c, str) and not c.strip()) for c in row)


def find_header_row(ws, max_scan=HEADER_SCAN_ROWS):
    """Row number (1-based) of the first row, among the first max_scan
    rows, that contains a recognizable Tag Number header. None if not
    found."""
    for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        for cell in row:
            if HEADER_ALIASES.get(normalize_header(cell)) == "tag":
                return row_idx
    return None


def build_column_map(header_row_values):
    """{canonical_key: 0-based index into the header row tuple}. First
    column wins on an accidental duplicate header."""
    col_map = {}
    for idx, cell in enumerate(header_row_values):
        key = HEADER_ALIASES.get(normalize_header(cell))
        if key and key not in col_map:
            col_map[key] = idx
    return col_map


def _iter_data_rows(ws, header_row_idx, blank_streak_stop=BLANK_STREAK_STOP):
    """Yields every row after the header, stopping after blank_streak_stop
    consecutive fully-blank rows (finding 1) - this is what turns a claimed
    1,048,575-row sheet into a sub-second scan of its real ~20 rows."""
    blanks = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if _is_blank_row(row):
            blanks += 1
            if blanks >= blank_streak_stop:
                return
            continue
        blanks = 0
        yield row


def parse_sheet(ws, sheet_name):
    """Returns (items, summary). items: a list of normalized dicts (see
    module docstring / master_list.json's shape in the spec) - everything
    except 'mapped_series', which the import step adds once the user has
    confirmed the sheet->series mapping. summary: counts + problems for
    the import dialog's per-sheet review table (10.4 step 1)."""
    header_row_idx = find_header_row(ws)
    summary = {
        "sheet_name": sheet_name, "rows_found": 0, "transmitter": 0, "valve": 0,
        "out_of_scope": 0, "blank_tag_skipped": 0, "duplicate_tags": 0,
        "header_row": header_row_idx, "columns_found": [], "problems": [],
    }
    if header_row_idx is None:
        summary["problems"].append(
            "No recognizable 'Tag Number' header found in the first "
            f"{HEADER_SCAN_ROWS} rows - sheet skipped.")
        return [], summary

    header_values = next(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    col_map = build_column_map(header_values)
    summary["columns_found"] = sorted(col_map.keys())

    missing_required = REQUIRED_KEYS - col_map.keys()
    if missing_required:
        summary["problems"].append(
            f"Missing required column(s): {', '.join(sorted(missing_required))} - sheet skipped.")
        return [], summary

    def get(row, key):
        idx = col_map.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    items = []
    seen_tags = set()
    blank_tag_skipped = 0
    duplicate_tags = 0
    counts = {"transmitter": 0, "valve": 0, "out_of_scope": 0}

    for row in _iter_data_rows(ws, header_row_idx):
        tag_raw = get(row, "tag")
        tag = canonical_tag(tag_raw)
        if not tag:
            blank_tag_skipped += 1
            continue
        if tag in seen_tags:
            duplicate_tags += 1
            continue  # first occurrence wins, per 10.2
        seen_tags.add(tag)

        tag_type = tag_type_of(tag)
        type_desc = str(get(row, "type_desc") or "").strip()
        kind = classify_kind(tag_type, type_desc)
        counts[kind] += 1

        items.append({
            "tag_raw": str(tag_raw).strip(),
            "tag": tag,
            "area": parse_area_code(tag),
            "tag_type": tag_type,
            "service": str(get(row, "service") or "").strip(),
            "type_desc": type_desc,
            "kind": kind,
            "pid": str(get(row, "pid") or "").strip(),
            "line_number": str(get(row, "line_number") or "").strip(),
            "module_id": str(get(row, "module_id") or "").strip(),
            "ewp": str(get(row, "ewp") or "").strip(),
            "model": format_model(get(row, "model")),
            "manufacturer": str(get(row, "manufacturer") or "").strip(),
            "install_details": str(get(row, "install_details") or "").strip(),
            "remarks": str(get(row, "remarks") or "").strip(),
            "ml_received": parse_yes(get(row, "received")),
            "ml_installed": parse_yes(get(row, "installed")),
            "source_sheet": sheet_name,
        })

    summary.update({
        "rows_found": len(items),
        "transmitter": counts["transmitter"],
        "valve": counts["valve"],
        "out_of_scope": counts["out_of_scope"],
        "blank_tag_skipped": blank_tag_skipped,
        "duplicate_tags": duplicate_tags,
    })
    return items, summary


def read_master_list_file(path):
    """Parses every sheet in the given .xlsx. openpyxl read_only=True
    throughout - never trusts ws.max_row (finding 1). Returns
    (items_by_sheet, summaries): items_by_sheet is {sheet_name: [item, ...]}
    in workbook order; summaries is the matching list of per-sheet summary
    dicts, same order."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        items_by_sheet = {}
        summaries = []
        for sheet_name in wb.sheetnames:
            items, summary = parse_sheet(wb[sheet_name], sheet_name)
            items_by_sheet[sheet_name] = items
            summaries.append(summary)
        return items_by_sheet, summaries
    finally:
        wb.close()


def guess_series_mapping(sheet_names, known_series_numbers):
    """{sheet_name: series_number or None} - a leading digit-run of the
    sheet name matched against registered series numbers, e.g.
    '29103-K1B Pad' -> 29103 when 29103 is registered. A sheet whose name
    doesn't start with a registered series number (finding 8's '9050 SUS'
    vs series 500) maps to None - left for the user to confirm/pick by
    hand in the import dialog, never auto-guessed further than an exact
    number match."""
    known = set(known_series_numbers)
    mapping = {}
    for name in sheet_names:
        m = re.match(r"\s*(\d+)", name)
        guess = int(m.group(1)) if m else None
        mapping[name] = guess if guess in known else None
    return mapping
