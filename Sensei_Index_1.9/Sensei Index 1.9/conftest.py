# -*- coding: utf-8 -*-
"""
Root pytest conftest. Two jobs, both of which must happen before anything
else in the test session touches Qt or the app modules:

1. Force the offscreen Qt platform plugin, so the test suite (and CI) never
   needs a real display. Must be set before PySide6.QtWidgets is imported
   anywhere - including by pytest-qt itself - so this is the very first
   thing this file does.
2. Put this directory (where gui_app.py / data_access.py / the schemas
   live) on sys.path. The app has no package structure - every module does
   a bare `import data_access` - and pytest's own rootdir insertion only
   adds tests/ (no __init__.py there), not this directory.
"""
import os
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import json
import sys
from pathlib import Path

import pytest

APP_DIR = Path(__file__).resolve().parent
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

import openpyxl  # noqa: E402  (must come after the sys.path fix-up above)


def _build_test_workbook(path, series_numbers):
    """A minimal workbook that satisfies data_access's expectations for
    each given series number: one 'Transmitter Log <n>' and one
    'Valve Log <n>' sheet, header row 3 (matching export_to_pdf.HEADER_ROW
    / export_valve_to_pdf.HEADER_ROW) with every schema field's label, no
    data rows. Real save_row()/read_index_rows() calls against this behave
    exactly like against the production workbook."""
    import transmitter_schema
    import valve_schema

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for n in series_numbers:
        for label_prefix, schema in (("Transmitter", transmitter_schema), ("Valve", valve_schema)):
            ws = wb.create_sheet(f"{label_prefix} Log {n}")
            for col, field in enumerate(schema.LOG_COLUMNS, start=1):
                ws.cell(row=3, column=col).value = field["label"]
    wb.save(path)


@pytest.fixture
def isolated_app_dir(tmp_path, monkeypatch):
    """Points every data_access sidecar/workbook path constant at a fresh
    tmp_path so tests never read or write the real, checked-in
    Equipment_Inspection_Tracker.xlsx / JSON sidecars. Yields (tmp_path, da)."""
    import data_access as da

    series_numbers = [100, 29103]
    workbook_path = tmp_path / "Equipment_Inspection_Tracker.xlsx"
    _build_test_workbook(workbook_path, series_numbers)

    registry = {
        "series": [
            {"number": n, "transmitter_sheet": f"Transmitter Log {n}", "valve_sheet": f"Valve Log {n}"}
            for n in series_numbers
        ]
    }
    config_path = tmp_path / "series_registry.json"
    config_path.write_text(json.dumps(registry), encoding="utf-8")

    monkeypatch.setattr(da, "HERE", tmp_path, raising=False)
    monkeypatch.setattr(da, "WORKBOOK_PATH", workbook_path, raising=False)
    monkeypatch.setattr(da, "CONFIG_PATH", config_path, raising=False)
    monkeypatch.setattr(da, "TEMP_DIR", tmp_path / "temp_previews", raising=False)
    monkeypatch.setattr(da, "ASSETS_DIR", tmp_path / "assets", raising=False)
    monkeypatch.setattr(da, "SETTINGS_PATH", tmp_path / "app_settings.json", raising=False)
    monkeypatch.setattr(da, "STATUS_PATH", tmp_path / "equipment_status.json", raising=False)
    monkeypatch.setattr(da, "DRAFTS_PATH", tmp_path / "wizard_draft.json", raising=False)
    monkeypatch.setattr(da, "MASTER_LIST_PATH", tmp_path / "master_list.json", raising=False)
    monkeypatch.setattr(da, "UI_STATE_PATH", tmp_path / "ui_state.json", raising=False)
    monkeypatch.setattr(da, "ACTIVITY_LOG_PATH", tmp_path / "activity_log.jsonl", raising=False)
    monkeypatch.setattr(da, "BACKUPS_DIR", tmp_path / "backups", raising=False)
    monkeypatch.setattr(da, "REPORTS_DIR", tmp_path / "reports", raising=False)
    monkeypatch.setattr(da, "CLEANED_DIR", tmp_path / "cleaned", raising=False)
    da.invalidate_workbook_cache()

    yield tmp_path, da

    da.invalidate_workbook_cache()


class FakeMainWindow:
    """A minimal stand-in for gui_app.MainWindow - lets a GUI test
    construct a page/dialog that needs *a* main_window without building a
    real one (which would pull in the whole sidebar tree, dashboard,
    etc). Every method is a harmless no-op; a test asserting on what got
    called should monkeypatch the specific method it cares about rather
    than growing this class further."""

    def __init__(self, undo_stack):
        self.undo_stack = undo_stack

    def statusBar(self):
        class _StatusBar:
            def showMessage(self, *a, **k):
                pass
        return _StatusBar()

    def refresh_sidebar_and_dashboard(self):
        pass

    def refresh_current_view(self):
        pass

    def show_index(self, *a, **k):
        pass

    def show_coverage(self):
        pass

    def open_populating_wizard(self):
        pass

    def open_datasheet_import(self):
        pass

    def open_master_list_import(self):
        pass


@pytest.fixture
def fake_main_window():
    import gui_app
    return FakeMainWindow(gui_app.UndoManager())
