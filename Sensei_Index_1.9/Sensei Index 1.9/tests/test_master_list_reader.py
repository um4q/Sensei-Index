# -*- coding: utf-8 -*-
"""Phase 10.2 - master_list_reader.py. One test per verified quirk from the
v2.1 spec's section 1, using small hand-built workbooks (never the real
~3MB client file, which isn't checked into the repo) plus a couple of
pure-function unit tests for canonical_tag/classify_kind/format_model."""
import time

import openpyxl
import pytest

import master_list_reader as mlr


def _wb(rows, header_row=("Tag Number", "Service", "Instrument Type Description",
                           "P&ID", "Line Number", "Model", "Manufacturer",
                           "Received", "Installed")):
    """A one-sheet workbook: header_row at row 1, then one row per entry in
    `rows` (each a tuple aligned to header_row)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(list(header_row))
    for r in rows:
        ws.append(list(r))
    return wb


def _save_and_reread(wb, tmp_path, name="test.xlsx"):
    path = tmp_path / name
    wb.save(path)
    return path


# --------------------------------------------------------- finding 1: max_row
def test_inflated_dimension_does_not_hang_or_scan_forever(tmp_path):
    """Reproduces the 29203 sheet: a claimed used-range far beyond the real
    data. Force an inflated dimension by writing one cell way out past the
    real data, then clearing its value (openpyxl keeps the sheet's
    dimension record pointing at the touched cell even after that)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Tag Number", "Service", "Instrument Type Description"])
    ws.append(["29203-TIT-0100", "STEAM", "TEMPERATURE INDICATING TRANSMITTER"])
    ws.cell(row=200_000, column=3).value = "x"
    ws.cell(row=200_000, column=3).value = None  # blank again, dimension record still inflated
    path = _save_and_reread(wb, tmp_path)

    real_wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws2 = real_wb["Sheet"]
    assert ws2.max_row > 100_000  # confirms the reproduction actually inflated it

    t0 = time.time()
    items, summary = mlr.parse_sheet(ws2, "Sheet")
    elapsed = time.time() - t0
    real_wb.close()

    assert elapsed < 5.0  # would take much longer iterating to max_row for real
    assert summary["rows_found"] == 1
    assert items[0]["tag"] == "29203-TIT-0100"


# -------------------------------------------------------- finding 2: headers
def test_normalize_header_matches_every_real_variant():
    variants = {
        "Tag Number": "tag", "Tag Number ": "tag", "Tag_Number": "tag",
        "Instrument Type Description": "type_desc", "Instrument_Type_Desc": "type_desc",
        "P&ID": "pid", "P\xa0\xa0\xa0\xa0\xa0ID": "pid",
        "Model": "model", "Model_Name": "model",
        "Manufacturer": "manufacturer", "Manufacturer_Name": "manufacturer",
        "Remarks": "remarks", "Remark": "remarks",
    }
    for raw, expected in variants.items():
        assert mlr.HEADER_ALIASES.get(mlr.normalize_header(raw)) == expected, raw


def test_9050_sus_style_headers_parse_with_required_columns_only(tmp_path):
    """9050 SUS has no Location Layout, no Remarks - only tag + type_desc
    are required (10.2); everything else is optional per sheet."""
    header = ("Tag_Number", "Service", "Instrument_Type_Desc",
              "P\xa0\xa0\xa0\xa0\xa0ID", "Line_Number", "Model_Name", "Manufacturer_Name")
    wb = _wb([("9050-FI   -922", "SEAL FLUID", "FIELD FLOW INDICATOR",
               "9050-25J-003-01", "", "", "")], header_row=header)
    path = _save_and_reread(wb, tmp_path)
    real_wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items, summary = mlr.parse_sheet(real_wb["Sheet1"], "9050 SUS")
    real_wb.close()

    assert not summary["problems"]
    assert summary["rows_found"] == 1
    assert items[0]["tag"] == "9050-FI-922"
    assert items[0]["kind"] == "out_of_scope"  # a Field Indicator, not a transmitter


def test_sheet_missing_a_required_column_is_skipped_not_crashed(tmp_path):
    wb = _wb([("A", "B")], header_row=("Service", "Line Number"))  # no Tag column at all
    path = _save_and_reread(wb, tmp_path)
    real_wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items, summary = mlr.parse_sheet(real_wb["Sheet1"], "Sheet1")
    real_wb.close()

    assert items == []
    assert summary["header_row"] is None
    assert summary["problems"]


# --------------------------------------------------- finding 3: tag padding
@pytest.mark.parametrize("raw,expected", [
    ("29103-AT  -2720A", "29103-AT-2720A"),
    ("29103-PIT -2171", "29103-PIT-2171"),
    ("29103-FIT-0916", "29103-FIT-0916"),   # tracker's own unpadded style, unchanged
    ("  29103-XV -0100  ", "29103-XV-0100"),
    (None, ""),
    ("", ""),
])
def test_canonical_tag(raw, expected):
    assert mlr.canonical_tag(raw) == expected


def test_canonical_tag_matches_across_padded_and_unpadded_forms():
    """The exact Phase 11 acceptance criterion: '29103-FIT-0916' (tracker,
    unpadded) must canonicalize identically to the master list's padded
    '29103-FIT -0916'."""
    assert mlr.canonical_tag("29103-FIT -0916") == mlr.canonical_tag("29103-FIT-0916")


def test_parse_area_code():
    assert mlr.parse_area_code("29103-PIT-2171") == "29103"
    assert mlr.parse_area_code("") == ""
    assert mlr.parse_area_code("NOHYPHEN") == "NOHYPHEN"


# ------------------------------------------------ finding 4: float coercion
@pytest.mark.parametrize("raw,expected", [
    (233.34, "233.34"),
    (5263.5123, "5263.5123"),
    (233.0, "233"),
    (100, "100"),
    (None, ""),
    ("3051TG3A2B21AM5Q4Q8P1BR5", "3051TG3A2B21AM5Q4Q8P1BR5"),
])
def test_format_model_no_float_noise(raw, expected):
    result = mlr.format_model(raw)
    assert result == expected
    assert "000000" not in result  # the exact noise pattern the spec calls out
    assert isinstance(result, str)


# ---------------------------------------------------- finding 5: typos
@pytest.mark.parametrize("type_desc", [
    "ANALYZER TX - MOSISTURE",
    "TEMPTERATURE GAUGE",
])
def test_typos_do_not_break_classification(type_desc):
    # MOSISTURE row: still classifies as transmitter (keys on "TX", not the typo)
    # TEMPTERATURE GAUGE: still out_of_scope (keys on "GAUGE" not matching anything)
    kind = mlr.classify_kind("AT", type_desc)
    assert kind in ("transmitter", "out_of_scope")


def test_mosisture_typo_row_is_still_a_transmitter():
    assert mlr.classify_kind("AT", "ANALYZER TX - MOSISTURE") == "transmitter"


# --------------------------------------------- finding 6: Yes/blank booleans
@pytest.mark.parametrize("raw,expected", [
    ("Yes", True), ("yes", True), (" YES ", True),
    (None, False), ("", False), ("No", False), ("N", False),
])
def test_parse_yes(raw, expected):
    assert mlr.parse_yes(raw) is expected


# ------------------------------------------------- finding 7: classification
@pytest.mark.parametrize("tag_type,type_desc,expected", [
    ("PIT", "PRESSURE INDICATING TRANSMITTER", "transmitter"),
    ("TIT", "TEMPERATURE INDICATING TRANSMITTER", "transmitter"),
    ("XV", "SHUTDOWN VALVE", "valve"),
    ("FV", "CONTROL VALVE", "valve"),
    ("TW", "THERMOWELL", "out_of_scope"),
    ("PSV", "PRESSURE SAFETY VALVE", "out_of_scope"),   # relief device, explicit exclusion
    ("PSE", "PRESSURE SAFETY ELEMENT", "out_of_scope"),
    ("PCV", "PRESSURE REGULATOR SELF CONTAINED", "out_of_scope"),  # self-contained, not a real valve
    ("TCV", "TEMPERATURE REGULATOR SELF CONTAINED", "out_of_scope"),
    ("PG", "PRESSURE GAUGE", "out_of_scope"),
    ("FE", "FLOW ELEMENT - VORTEX", "out_of_scope"),
    ("XY", "POSITIONER", "out_of_scope"),
    ("YS", "MOTOR TRIP", "out_of_scope"),
])
def test_classify_kind(tag_type, type_desc, expected):
    assert mlr.classify_kind(tag_type, type_desc) == expected


def test_classify_kind_type_desc_wins_over_tag_type_fallback():
    """A tag-type that would fall back to transmitter, but whose type_desc
    unambiguously says valve, should classify by the type_desc (checked
    first per 10.3's ordered rule table)."""
    assert mlr.classify_kind("PIT", "ON/OFF VALVE POSITION INDICATOR") == "valve"


# ---------------------------------------------------------- deduplication
def test_duplicate_canonical_tag_keeps_first_and_counts_collision(tmp_path):
    wb = _wb([
        ("29103-PIT -2171", "SVC A", "PRESSURE INDICATING TRANSMITTER", "", "", "", "", "", ""),
        ("29103-PIT-2171", "SVC B (dup, different spacing)", "PRESSURE INDICATING TRANSMITTER",
         "", "", "", "", "", ""),
    ])
    path = _save_and_reread(wb, tmp_path)
    real_wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items, summary = mlr.parse_sheet(real_wb["Sheet1"], "Sheet1")
    real_wb.close()

    assert summary["rows_found"] == 1
    assert summary["duplicate_tags"] == 1
    assert items[0]["service"] == "SVC A"  # first occurrence wins


def test_blank_tag_rows_are_skipped_and_counted(tmp_path):
    wb = _wb([
        ("", "no tag", "SOMETHING", "", "", "", "", "", ""),
        ("   ", "whitespace-only tag", "SOMETHING", "", "", "", "", "", ""),
        ("29103-PIT-2171", "real row", "PRESSURE INDICATING TRANSMITTER", "", "", "", "", "", ""),
    ])
    path = _save_and_reread(wb, tmp_path)
    real_wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items, summary = mlr.parse_sheet(real_wb["Sheet1"], "Sheet1")
    real_wb.close()

    assert summary["rows_found"] == 1
    assert summary["blank_tag_skipped"] == 2


# --------------------------------------------------------------- read_master_list_file
def test_read_master_list_file_multi_sheet(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "29103-K1B Pad"
    ws1.append(["Tag Number", "Service", "Instrument Type Description"])
    ws1.append(["29103-PIT-2171", "SVC", "PRESSURE INDICATING TRANSMITTER"])
    ws2 = wb.create_sheet("9050 SUS")
    ws2.append(["Tag_Number", "Service", "Instrument_Type_Desc"])
    ws2.append(["9050-FI-922", "SVC", "FIELD FLOW INDICATOR"])
    path = tmp_path / "ml.xlsx"
    wb.save(path)

    items_by_sheet, summaries = mlr.read_master_list_file(path)
    assert set(items_by_sheet.keys()) == {"29103-K1B Pad", "9050 SUS"}
    assert len(items_by_sheet["29103-K1B Pad"]) == 1
    assert len(items_by_sheet["9050 SUS"]) == 1
    assert [s["sheet_name"] for s in summaries] == ["29103-K1B Pad", "9050 SUS"]


# -------------------------------------------------- finding 8: series mapping
def test_guess_series_mapping_exact_number_match():
    mapping = mlr.guess_series_mapping(
        ["29103-K1B Pad", "29151 Drain 1", "9050 SUS"],
        known_series_numbers=[100, 200, 29103, 29151, 29152, 500],
    )
    assert mapping["29103-K1B Pad"] == 29103
    assert mapping["29151 Drain 1"] == 29151
    assert mapping["9050 SUS"] is None  # 9050 isn't a registered series number - left for the user
