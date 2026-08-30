# -*- coding: utf-8 -*-
"""Phase 17 - Client progress report export."""
import openpyxl


def _ml_item(tag, kind, **overrides):
    item = {
        "tag_raw": tag, "tag": tag, "area": tag.split("-")[0], "tag_type": tag.split("-")[1],
        "service": "SVC", "type_desc": "PRESSURE INDICATING TRANSMITTER", "kind": kind, "pid": "",
        "line_number": "", "module_id": "", "ewp": "", "model": "", "manufacturer": "",
        "install_details": "", "remarks": "", "ml_received": False, "ml_installed": False,
        "source_sheet": "29103-K1B Pad", "mapped_series": 29103,
    }
    item.update(overrides)
    return item


def _seed_master_list(da, items):
    da.save_master_list({
        "imported_at": "2026-08-30T00:00:00", "source_file": "x.xlsx", "source_mtime": 0,
        "sheet_series_map": {}, "items": items,
    })


def test_no_master_list_degrades_to_tracker_only(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171", "serial_number": "SN-1"})

    data = da.build_progress_report_data()
    assert data["has_master_list"] is False
    assert len(data["summary"]) == 1
    assert data["summary"][0]["total_in_scope"] == 1
    assert data["missing"] == []


def test_summary_numbers_match_coverage_and_dashboard(isolated_app_dir):
    """17's own acceptance criterion: report numbers equal what the app's
    own screens show at that moment, for the same buckets."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171", "serial_number": "SN-1"})
    da.set_status(29103, "transmitter", "29103-PIT-2171", installed=True)
    _seed_master_list(da, [_ml_item("29103-PIT-2171", "transmitter")])

    report = da.build_progress_report_data()
    coverage = da.reconcile_master_list()

    row_summary = next(r for r in report["summary"] if r["kind"] == "transmitter")
    assert row_summary["total_in_scope"] == len(coverage["matched"]) + len(coverage["missing"])
    assert row_summary["installed"] == sum(
        1 for p in coverage["matched"]
        if next(m for m in p["progress"]["milestones"] if m["id"] == "installed")["done"])

    dashboard_avg = da.series_progress_summary(29103)["transmitter"]["avg_percent"]
    # single matched row -> report's avg for this area/kind matches Dashboard's
    # series-level average too (same one row driving both).
    assert row_summary["avg_percent"] == dashboard_avg


def test_missing_sheet_lists_unmatched_master_items(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    _seed_master_list(da, [_ml_item("29103-PIT-2171", "transmitter", service="MAINLINE SVC")])
    data = da.build_progress_report_data()
    assert len(data["missing"]) == 1
    assert data["missing"][0]["tag"] == "29103-PIT-2171"
    assert data["missing"][0]["service"] == "MAINLINE SVC"


def test_flags_sheet_includes_installed_mismatch(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})
    _seed_master_list(da, [_ml_item("29103-PIT-2171", "transmitter", ml_installed=True)])

    data = da.build_progress_report_data()
    assert any("Installed mismatch" in f["detail"] for f in data["flags"])


def test_flags_sheet_includes_tag_shape_warning(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "not a valid shape"})
    data = da.build_progress_report_data()
    assert any(f["tag"] == "not a valid shape" for f in data["flags"])


def test_flags_sheet_includes_duplicate_serials(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    r1 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", r1, {"tag": "29103-PIT-0001", "serial_number": "SN-DUP"})
    r2 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", r2, {"tag": "29103-PIT-0002", "serial_number": "SN-DUP"})

    data = da.build_progress_report_data()
    assert any("Duplicate serial" in f["detail"] for f in data["flags"])


def test_write_progress_report_xlsx_produces_a_real_workbook(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})
    _seed_master_list(da, [_ml_item("29103-PIT-2171", "transmitter")])

    data = da.build_progress_report_data()
    out_path = tmp_path / "report.xlsx"
    da.write_progress_report_xlsx(data, out_path)

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == ["Summary", "Missing", "Flags"]
    ws1 = wb["Summary"]
    assert ws1.cell(row=1, column=1).value == "Area"
    assert ws1.freeze_panes == "A2"
    assert ws1.auto_filter.ref is not None
    assert ws1.cell(row=1, column=1).font.bold is True


def test_write_progress_report_xlsx_tracker_only_note(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})

    data = da.build_progress_report_data()
    out_path = tmp_path / "report.xlsx"
    da.write_progress_report_xlsx(data, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws1 = wb["Summary"]
    assert "No master list" in ws1.cell(row=1, column=1).value
    assert ws1.cell(row=3, column=1).value == "Area"  # header shifted down past the note
    assert ws1.freeze_panes == "A4"


def test_export_progress_report_writes_into_reports_dir(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})

    out_path = da.export_progress_report()
    assert out_path.exists()
    assert out_path.parent == da.REPORTS_DIR
    assert out_path.name.startswith("Progress_Report_")


def test_export_progress_report_does_not_clobber_same_day_export(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})

    first = da.export_progress_report()
    second = da.export_progress_report()
    assert first != second
    assert first.exists() and second.exists()


def test_report_works_with_zero_rows_anywhere(isolated_app_dir):
    """No crash, no divide-by-zero, on a totally empty tracker."""
    tmp_path, da = isolated_app_dir
    data = da.build_progress_report_data()
    assert data["summary"] == []
    assert data["missing"] == []
    out_path = tmp_path / "empty.xlsx"
    da.write_progress_report_xlsx(data, out_path)
    assert out_path.exists()
