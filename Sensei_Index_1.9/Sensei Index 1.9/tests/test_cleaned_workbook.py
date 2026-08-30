# -*- coding: utf-8 -*-
"""Phase 18 - "Cleaned copy" export: auto-fix safe whitespace/casing
issues, flag (never touch) duplicate serials and malformed tag shapes,
never insert/delete/reorder a row, never open the live workbook for
writing."""
import openpyxl
import pytest

import export_to_pdf


def test_already_clean_row_produces_no_fixes_or_flags(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171", "serial_number": "SN-001"})

    plan = da.build_cleanup_plan()
    assert plan["fixes"] == []
    assert plan["flags"] == []


def test_trims_whitespace_and_reports_as_fix(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171", "serial_number": "  SN-001  "})

    plan = da.build_cleanup_plan()
    fix = next(f for f in plan["fixes"] if f["field"] == "serial_number")
    assert fix["old"] == "  SN-001  "
    assert fix["new"] == "SN-001"
    assert fix["row"] == row
    assert fix["series"] == 29103
    assert fix["key_value"] == "29103-PIT-2171"


def test_uppercases_key_field_and_reports_as_fix(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-pit-2171"})

    plan = da.build_cleanup_plan()
    fix = next(f for f in plan["fixes"] if f["field"] == "tag")
    assert fix["old"] == "29103-pit-2171"
    assert fix["new"] == "29103-PIT-2171"


def test_clean_cell_is_never_touched_even_if_field_map_present(isolated_app_dir):
    """A cell that's already exactly right (no whitespace, key already
    upper) must not show up as a 'fix' at all - only an actual change is
    reported."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171", "system_number": "SYS-1"})

    plan = da.build_cleanup_plan()
    assert plan["fixes"] == []


def test_date_field_never_touched_even_with_stray_whitespace(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171", "te1_caldate": "  2024-01-01  "})

    plan = da.build_cleanup_plan()
    assert not any(f["field"] == "te1_caldate" for f in plan["fixes"])


def test_non_string_cell_type_never_touched(isolated_app_dir):
    """A cell holding a raw non-str value (e.g. a number typed straight
    into Excel) is never rewritten by the whitespace/case fixer - only
    plain str cells are eligible."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171", "serial_number": "12345"})

    wb = openpyxl.load_workbook(da.WORKBOOK_PATH)
    ws = wb["Transmitter Log 29103"]
    col = export_to_pdf.load_column_map(ws)["serial_number"]
    ws.cell(row=row, column=col).value = 12345  # raw int, not str
    wb.save(da.WORKBOOK_PATH)
    da.invalidate_workbook_cache()

    plan = da.build_cleanup_plan()
    assert not any(f["row"] == row and f["field"] == "serial_number" for f in plan["fixes"])


def test_flags_tag_that_still_looks_wrong_after_normalizing(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "not a valid shape"})

    plan = da.build_cleanup_plan()
    flag = next(f for f in plan["flags"] if f["row"] == row)
    assert flag["key_value"] == "NOT A VALID SHAPE"
    assert "doesn't look like AREA-TYPE-NUMBER" in flag["detail"]


def test_flags_duplicate_serials_without_touching_them(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    r1 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", r1, {"tag": "29103-PIT-0001", "serial_number": "SN-DUP"})
    r2 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", r2, {"tag": "29103-PIT-0002", "serial_number": "SN-DUP"})

    plan = da.build_cleanup_plan()
    dup_flags = [f for f in plan["flags"] if "Duplicate serial" in f["detail"]]
    assert {f["row"] for f in dup_flags} == {r1, r2}
    # never "fixed" - a duplicate serial isn't a whitespace/case problem
    assert not any(f["field"] == "serial_number" for f in plan["fixes"])


def test_blank_row_is_never_touched(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "  29103-PIT-2171  "})
    blank_row = row + 1  # never written to - still empty

    plan = da.build_cleanup_plan()
    assert not any(f["row"] == blank_row for f in plan["fixes"])
    assert not any(f["row"] == blank_row for f in plan["flags"])


def test_export_writes_new_file_and_never_touches_the_live_workbook(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-pit-2171", "serial_number": "  SN-001  "})

    before_bytes = da.WORKBOOK_PATH.read_bytes()
    out_path = da.export_cleaned_workbook()
    after_bytes = da.WORKBOOK_PATH.read_bytes()

    assert before_bytes == after_bytes  # the live workbook was only ever read
    assert out_path.exists()
    assert out_path.parent == da.CLEANED_DIR
    assert out_path.name.startswith("Equipment_Inspection_Tracker_CLEANED_")


def test_export_applies_fixes_and_highlights_them(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-pit-2171", "serial_number": "  SN-001  "})

    out_path = da.export_cleaned_workbook()
    out_wb = openpyxl.load_workbook(out_path)
    ws = out_wb["Transmitter Log 29103"]
    col_map = export_to_pdf.load_column_map(ws)

    assert ws.cell(row=row, column=col_map["tag"]).value == "29103-PIT-2171"
    assert ws.cell(row=row, column=col_map["serial_number"]).value == "SN-001"
    assert "FFEB9C" in ws.cell(row=row, column=col_map["tag"]).fill.fgColor.rgb
    assert "FFEB9C" in ws.cell(row=row, column=col_map["serial_number"]).fill.fgColor.rgb


def test_export_highlights_flagged_rows_key_cell(isolated_app_dir):
    """A tag that's flagged but needs NO fix (already upper-case, no
    whitespace) - pure flag-only case, gets the plain red flag color."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "NOTAVALIDSHAPE"})

    out_path = da.export_cleaned_workbook()
    out_wb = openpyxl.load_workbook(out_path)
    ws = out_wb["Transmitter Log 29103"]
    col_map = export_to_pdf.load_column_map(ws)
    assert "FFC7CE" in ws.cell(row=row, column=col_map["tag"]).fill.fgColor.rgb


def test_export_highlights_both_fixed_and_flagged_cell_distinctly(isolated_app_dir):
    """A tag that needed a fix (case) AND is still flagged afterward (bad
    shape) must not silently lose one signal to the other - it gets its
    own distinct color, and the value is still correctly written."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "not a valid shape"})

    out_path = da.export_cleaned_workbook()
    out_wb = openpyxl.load_workbook(out_path)
    ws = out_wb["Transmitter Log 29103"]
    col_map = export_to_pdf.load_column_map(ws)
    cell = ws.cell(row=row, column=col_map["tag"])
    assert cell.value == "NOT A VALID SHAPE"
    assert "FFD966" in cell.fill.fgColor.rgb
    assert "FFEB9C" not in cell.fill.fgColor.rgb
    assert "FFC7CE" not in cell.fill.fgColor.rgb


def test_export_appends_cleanup_log_sheet_after_original_sheets(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    original_names = list(openpyxl.load_workbook(da.WORKBOOK_PATH).sheetnames)

    out_path = da.export_cleaned_workbook()
    out_wb = openpyxl.load_workbook(out_path)
    assert out_wb.sheetnames == original_names + ["Cleanup Log"]


def test_cleanup_log_notes_when_nothing_to_fix_or_flag(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})

    out_path = da.export_cleaned_workbook()
    ws = openpyxl.load_workbook(out_path)["Cleanup Log"]
    values = [c.value for r in ws.iter_rows() for c in r if c.value]
    assert any("already clean" in str(v) for v in values)


def test_cleanup_log_lists_fix_and_flag_details(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-pit-2171", "serial_number": "  SN-001  "})
    row2 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row2, {"tag": "not a valid shape"})

    out_path = da.export_cleaned_workbook()
    ws = openpyxl.load_workbook(out_path)["Cleanup Log"]
    values = [c.value for r in ws.iter_rows() for c in r if c.value]
    assert any(v == "SN-001" for v in values)
    assert any("AREA-TYPE-NUMBER" in str(v) for v in values)


def test_export_never_inserts_or_deletes_rows(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-pit-2171"})

    before_max_row = openpyxl.load_workbook(da.WORKBOOK_PATH)["Transmitter Log 29103"].max_row
    out_path = da.export_cleaned_workbook()
    after_max_row = openpyxl.load_workbook(out_path)["Transmitter Log 29103"].max_row
    assert before_max_row == after_max_row


def test_export_does_not_clobber_same_day_export(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})

    first = da.export_cleaned_workbook()
    second = da.export_cleaned_workbook()
    assert first != second
    assert first.exists() and second.exists()


def test_export_raises_if_no_workbook_yet(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    da.WORKBOOK_PATH.unlink()
    with pytest.raises(FileNotFoundError):
        da.export_cleaned_workbook()


def test_export_works_with_zero_rows_anywhere(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    out_path = da.export_cleaned_workbook()
    assert out_path.exists()
    ws = openpyxl.load_workbook(out_path)["Cleanup Log"]
    values = [c.value for r in ws.iter_rows() for c in r if c.value]
    assert any("already clean" in str(v) for v in values)


def test_formula_cell_is_never_touched_even_though_its_raw_value_is_a_str(isolated_app_dir):
    """A formula's raw value (data_only=False) is a plain str too (its
    formula TEXT, e.g. '=CONCATENATE(...)') - rewriting that would
    silently mutate a string literal embedded inside the formula, changing
    what it actually computes. Must be excluded exactly like a
    numeric/bool/date cell is."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})

    wb = openpyxl.load_workbook(da.WORKBOOK_PATH)
    ws = wb["Transmitter Log 29103"]
    col = export_to_pdf.load_column_map(ws)["system_number"]
    formula = '=CONCATENATE("sys","-",1)'
    ws.cell(row=row, column=col).value = formula
    wb.save(da.WORKBOOK_PATH)
    da.invalidate_workbook_cache()

    plan = da.build_cleanup_plan()
    assert not any(f["row"] == row and f["field"] == "system_number" for f in plan["fixes"])

    out_path = da.export_cleaned_workbook()
    out_ws = openpyxl.load_workbook(out_path)["Transmitter Log 29103"]
    assert out_ws.cell(row=row, column=col).value == formula


def test_formula_in_key_field_is_flagged_but_never_rewritten(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")

    wb = openpyxl.load_workbook(da.WORKBOOK_PATH)
    ws = wb["Transmitter Log 29103"]
    col = export_to_pdf.load_column_map(ws)["tag"]
    formula = '=CONCATENATE("29103","-pit-","2171")'
    ws.cell(row=row, column=col).value = formula
    wb.save(da.WORKBOOK_PATH)
    da.invalidate_workbook_cache()

    plan = da.build_cleanup_plan()
    assert not any(f["row"] == row and f["field"] == "tag" for f in plan["fixes"])

    out_path = da.export_cleaned_workbook()
    out_ws = openpyxl.load_workbook(out_path)["Transmitter Log 29103"]
    assert out_ws.cell(row=row, column=col).value == formula  # untouched, still a live formula


def test_whitespace_only_key_field_still_counts_as_a_used_row(isolated_app_dir):
    """Matches read_index_rows()/find_first_blank_row()'s own definition
    of "used" (raw value not in (None, "")) - a whitespace-only tag is a
    real row there, so it must be one here too: its OTHER fields still get
    fixed, not silently skipped."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "   ", "system_number": "  SYS-9  "})

    assert any(r["row"] == row for r in da.read_index_rows(29103, "transmitter"))

    plan = da.build_cleanup_plan()
    fix = next(f for f in plan["fixes"] if f["row"] == row and f["field"] == "system_number")
    assert fix["old"] == "  SYS-9  "
    assert fix["new"] == "SYS-9"


def test_cleanup_log_name_collision_gets_a_distinct_suffixed_name(isolated_app_dir):
    """If the live workbook already carries a sheet literally named
    'Cleanup Log' (e.g. a previous cleaned copy was promoted to become the
    new live tracker), this run's real log must never be silently written
    into - or on top of - that old one under openpyxl's own silent
    auto-rename behavior. The old sheet is left exactly as it was, and
    this run's log gets its own clearly distinct name."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-pit-2171"})

    wb = openpyxl.load_workbook(da.WORKBOOK_PATH)
    stale = wb.create_sheet("Cleanup Log")
    stale["A1"] = "STALE - from an earlier cleanup run"
    wb.save(da.WORKBOOK_PATH)
    da.invalidate_workbook_cache()

    out_path = da.export_cleaned_workbook()
    out_wb = openpyxl.load_workbook(out_path)
    assert out_wb["Cleanup Log"]["A1"].value == "STALE - from an earlier cleanup run"
    assert "Cleanup Log (2)" in out_wb.sheetnames
    new_log_values = [c.value for r in out_wb["Cleanup Log (2)"].iter_rows() for c in r if c.value]
    assert any(v == "29103-PIT-2171" for v in new_log_values)


def test_duplicate_serial_flag_key_value_is_normalized_like_every_other_entry(isolated_app_dir):
    """A duplicate-serial flag's key_value/detail text must match the
    same trimmed+upper-cased casing every other Cleanup Log entry (and the
    actual fixed cell, if this row's key also needed a fix) uses - not the
    raw, un-normalized value find_all_duplicate_serials() itself returns
    (which Phase 17's Progress Report still relies on as-is)."""
    tmp_path, da = isolated_app_dir
    r1 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", r1, {"tag": "  29103-pit-0001  ", "serial_number": "SN-DUP"})
    r2 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", r2, {"tag": "29103-PIT-0002", "serial_number": "SN-DUP"})

    plan = da.build_cleanup_plan()
    fix = next(f for f in plan["fixes"] if f["row"] == r1 and f["field"] == "tag")
    assert fix["new"] == "29103-PIT-0001"

    flag = next(f for f in plan["flags"] if f["row"] == r1 and "Duplicate serial" in f["detail"])
    assert flag["key_value"] == "29103-PIT-0001"  # matches the fix, not the raw lowercase original
    assert "29103-PIT-0001" in flag["detail"]
    assert "29103-pit-0001" not in flag["detail"]


def test_valve_rows_are_fixed_and_flagged_same_as_transmitter_rows(isolated_app_dir):
    """The valve branch uses a different export module, key_field
    ("equip_number"), and serial field ("valve_serial") than transmitter -
    exercised here explicitly rather than only ever reached with empty
    sheets."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "valve")
    da.save_row(29103, "valve", row, {"equip_number": "29103-pv-0001", "valve_serial": "  VS-001  "})

    plan = da.build_cleanup_plan()
    key_fix = next(f for f in plan["fixes"] if f["row"] == row and f["field"] == "equip_number")
    assert key_fix["new"] == "29103-PV-0001"
    serial_fix = next(f for f in plan["fixes"] if f["row"] == row and f["field"] == "valve_serial")
    assert serial_fix["old"] == "  VS-001  "
    assert serial_fix["new"] == "VS-001"

    out_path = da.export_cleaned_workbook()
    out_ws = openpyxl.load_workbook(out_path)["Valve Log 29103"]
    import export_valve_to_pdf
    col_map = export_valve_to_pdf.load_column_map(out_ws)
    assert out_ws.cell(row=row, column=col_map["equip_number"]).value == "29103-PV-0001"
    assert out_ws.cell(row=row, column=col_map["valve_serial"]).value == "VS-001"


def test_valve_duplicate_serials_are_flagged(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    r1 = da.find_first_blank_row(29103, "valve")
    da.save_row(29103, "valve", r1, {"equip_number": "29103-PV-0001", "valve_serial": "VS-DUP"})
    r2 = da.find_first_blank_row(29103, "valve")
    da.save_row(29103, "valve", r2, {"equip_number": "29103-PV-0002", "valve_serial": "VS-DUP"})

    plan = da.build_cleanup_plan()
    dup_flags = [f for f in plan["flags"] if f["equip_key"] == "valve" and "Duplicate serial" in f["detail"]]
    assert {f["row"] for f in dup_flags} == {r1, r2}
