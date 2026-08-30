# -*- coding: utf-8 -*-
"""Phase 16.1 (automatic backups) and 16.2 (Excel-lock detection)."""
import time

import pytest


def test_first_save_creates_a_backup(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    assert da.list_backups() == []
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})
    backups = da.list_backups()
    assert len(backups) == 1
    assert backups[0]["name"].startswith("Equipment_Inspection_Tracker.")


def test_backup_is_a_snapshot_of_the_workbook_before_the_write(isolated_app_dir):
    """The backup must capture the PRE-write state, not the post-write
    one - otherwise it's useless as a safety net."""
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "BEFORE"})

    row2 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row2, {"tag": "AFTER"})  # triggers a 2nd backup? see interval test

    # force a backup on this 3rd save by making the "last backup" look old
    backups = da.list_backups()
    old_time = time.time() - 3600
    import os
    os.utime(backups[0]["path"], (old_time, old_time))

    row3 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row3, {"tag": "THIRD"})

    newest_backup = da.list_backups()[0]["path"]
    import openpyxl
    wb = openpyxl.load_workbook(newest_backup, read_only=True)
    ws = wb["Transmitter Log 29103"]
    tags = {ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)}
    wb.close()
    # the backup was taken BEFORE row3's write - "THIRD" shouldn't be in it,
    # but BEFORE/AFTER (already-saved before this backup) should be.
    assert "THIRD" not in tags


def test_backup_not_repeated_within_the_interval(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-0001"})
    assert len(da.list_backups()) == 1

    row2 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row2, {"tag": "29103-PIT-0002"})
    assert len(da.list_backups()) == 1  # still just one - within backup_interval_minutes


def test_backup_interval_setting_is_respected(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    da.set_setting("backup_interval_minutes", 0)  # "always back up"
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-0001"})
    row2 = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row2, {"tag": "29103-PIT-0002"})
    assert len(da.list_backups()) == 2


def test_backup_keep_prunes_oldest(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    da.set_setting("backup_interval_minutes", 0)
    da.set_setting("backup_keep", 3)
    for i in range(6):
        row = da.find_first_blank_row(29103, "transmitter")
        da.save_row(29103, "transmitter", row, {"tag": f"29103-PIT-{i:04d}"})
        time.sleep(0.01)  # ensure distinct mtimes/timestamps
    backups = da.list_backups()
    assert len(backups) == 3  # pruned down to backup_keep


def test_backup_now_ignores_interval(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-0001"})
    assert len(da.list_backups()) == 1
    da.backup_now()  # immediate, regardless of interval
    assert len(da.list_backups()) == 2


def test_backup_now_with_no_workbook_raises_cleanly(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    da.WORKBOOK_PATH.unlink()
    with pytest.raises(FileNotFoundError):
        da.backup_now()


def test_restore_backup_copies_snapshot_over_live_file_with_safety_snapshot(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "ORIGINAL"})
    old_backup = da.list_backups()[0]["path"]  # captured before "ORIGINAL" in this case is fine

    row2 = da.find_first_blank_row(29103, "transmitter")
    da.set_setting("backup_interval_minutes", 0)
    da.save_row(29103, "transmitter", row2, {"tag": "NEWER"})
    backups_before_restore = len(da.list_backups())

    safety = da.restore_backup(old_backup)
    assert safety is not None  # a safety snapshot of the pre-restore state was taken
    assert len(da.list_backups()) == backups_before_restore + 1

    tags = {r["tag"] for r in da.read_index_rows(29103, "transmitter")}
    assert "NEWER" not in tags  # the restore actually took effect


def test_restore_backup_missing_file_raises_cleanly(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    with pytest.raises(FileNotFoundError):
        da.restore_backup(tmp_path / "backups" / "nonexistent.xlsx")


def test_list_backups_empty_when_none_taken_yet(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    assert da.list_backups() == []


def test_backup_failure_never_blocks_the_actual_save(isolated_app_dir, monkeypatch, capsys):
    tmp_path, da = isolated_app_dir

    def boom():
        raise OSError("disk full")

    monkeypatch.setattr(da, "_write_backup_snapshot", boom)
    row = da.find_first_blank_row(29103, "transmitter")
    da.save_row(29103, "transmitter", row, {"tag": "29103-PIT-2171"})  # must not raise

    rows = da.read_index_rows(29103, "transmitter")
    assert any(r["tag"] == "29103-PIT-2171" for r in rows)
    assert "WARNING" in capsys.readouterr().out


# ------------------------------------------------------------- lock detection
def test_not_locked_when_no_lock_file(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    assert da.is_workbook_locked() is False


def test_locked_when_excel_lock_file_present(isolated_app_dir):
    tmp_path, da = isolated_app_dir
    lock_path = da.WORKBOOK_PATH.with_name(f"~${da.WORKBOOK_PATH.name}")
    lock_path.write_bytes(b"")
    assert da.is_workbook_locked() is True


def test_save_raises_workbook_locked_error_when_permission_error_and_locked(isolated_app_dir, monkeypatch):
    tmp_path, da = isolated_app_dir
    lock_path = da.WORKBOOK_PATH.with_name(f"~${da.WORKBOOK_PATH.name}")
    lock_path.write_bytes(b"")

    def boom(self, path):
        raise PermissionError("simulated")

    import openpyxl
    monkeypatch.setattr(openpyxl.Workbook, "save", boom)

    with pytest.raises(da.WorkbookLockedError, match="open in Excel"):
        row = da.find_first_blank_row(29103, "transmitter")
        da.save_row(29103, "transmitter", row, {"tag": "X"})


def test_save_raises_plain_permission_error_when_not_locked(isolated_app_dir, monkeypatch):
    """A genuine permissions problem (not Excel holding it open) still
    gets the original Program-Files/admin-rights message, not the
    Excel-specific one."""
    tmp_path, da = isolated_app_dir

    def boom(self, path):
        raise PermissionError("simulated")

    import openpyxl
    monkeypatch.setattr(openpyxl.Workbook, "save", boom)

    with pytest.raises(PermissionError, match="administrator"):
        row = da.find_first_blank_row(29103, "transmitter")
        da.save_row(29103, "transmitter", row, {"tag": "X"})
