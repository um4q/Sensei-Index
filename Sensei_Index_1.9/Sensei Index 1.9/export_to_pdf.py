#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Turns one or more rows of Transmitter_Inspection_Log.xlsx into filled,
ready-to-print PDFs (one PDF per row), using
Transmitter_Inspection_Test_Record_TEMPLATE.pdf as the blank form.

BASIC USE (exports every row whose "Export to PDF (Y/N)" column is Y):
    python3 export_to_pdf.py Transmitter_Inspection_Log.xlsx

EXPORT SPECIFIC ROWS ONLY (Excel row numbers, overrides the Y/N flag):
    python3 export_to_pdf.py Transmitter_Inspection_Log.xlsx --rows 4,7,12

EXPORT EVERY FILLED-IN ROW REGARDLESS OF THE FLAG:
    python3 export_to_pdf.py Transmitter_Inspection_Log.xlsx --all

ALSO PRODUCE ONE COMBINED PDF OF EVERYTHING EXPORTED THIS RUN:
    python3 export_to_pdf.py Transmitter_Inspection_Log.xlsx --merge

FLATTEN THE OUTPUT (bakes the values into the page, no longer fillable):
    python3 export_to_pdf.py Transmitter_Inspection_Log.xlsx --flatten

Requires: pip install openpyxl pypdf reportlab
"""
import argparse
import datetime
import re
import sys
from pathlib import Path

import openpyxl
from pypdf import PdfReader, PdfWriter
from pypdf.generic import DictionaryObject, NameObject, TextStringObject, ArrayObject
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import letter
import io

from transmitter_schema import LOG_COLUMNS, by_id
from transmitter_field_map import FIELD_MAP, PROC_CHECKBOXES, CHECKBOX_ON, CHECKBOX_OFF, REMARKS_LINE_FIELDS

# Frozen-aware: when bundled into an .exe by PyInstaller, __file__ points
# inside a temp extraction folder that's wiped on exit - sys.executable's
# folder is the exe's real, persistent location.
HERE = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) \
    else Path(__file__).resolve().parent
DEFAULT_TEMPLATE = HERE / "Transmitter_Inspection_Test_Record_TEMPLATE.pdf"
DEFAULT_OUTPUT_DIR = HERE / "output_pdfs"
HEADER_ROW = 3          # row with column labels in the Log sheet (combined workbook has a nav-button row above it)
FIRST_DATA_ROW = 4      # first row that may contain real (or example) data
SHEET_NAME = "Transmitter Log"

# Yanda QA Representative signature stamp - placed on Part 8, page 2, just
# above the "Signature:" underline (x=41.9-305.9, underline at y~282).
SIGNATURE_IMAGE = HERE / "assets" / "yanda_qa_signature_transparent.png"
SIGNATURE_PAGE_INDEX = 1  # page 2 (0-indexed)
SIGNATURE_X = 95
SIGNATURE_Y = 284
SIGNATURE_W = 95
SIGNATURE_H = 95 * (90 / 458)  # preserve the source image's aspect ratio


def sanitize(text, fallback):
    text = (text or "").strip()
    text = re.sub(r"[^A-Za-z0-9\-_. ]+", "", text).strip()
    return text or fallback


def cell_to_str(value):
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_column_map(ws):
    """Map each schema field's label -> its column index, by reading the
    header row. This means the script still works even if columns get
    reordered or new ones are inserted, as long as the header text matches."""
    label_to_col = {}
    for col in range(1, ws.max_column + 1):
        label = ws.cell(row=HEADER_ROW, column=col).value
        if label:
            label_to_col[str(label).strip()] = col

    field_to_col = {}
    missing = []
    for field in LOG_COLUMNS:
        col = label_to_col.get(field["label"])
        if col is None:
            missing.append(field["label"])
        else:
            field_to_col[field["id"]] = col
    if missing:
        print("WARNING: these expected columns were not found in the Log sheet header "
              f"(row {HEADER_ROW}) and will be left blank on export:")
        for m in missing:
            print("   -", m)
    return field_to_col


def rows_to_export(ws, field_to_col, explicit_rows, export_all):
    if explicit_rows:
        return sorted(set(explicit_rows))

    export_col = field_to_col.get("export_flag")
    tag_col = field_to_col.get("tag")
    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        flag = cell_to_str(ws.cell(row=r, column=export_col).value).upper() if export_col else ""
        has_tag = tag_col and cell_to_str(ws.cell(row=r, column=tag_col).value)
        if export_all:
            if has_tag:
                rows.append(r)
        elif flag == "Y":
            rows.append(r)
    return rows


def build_values_for_row(ws, field_to_col, row_num):
    """Returns a dict of REAL PDF field name -> value, using field_map.py to
    translate from schema/Excel field ids to the actual field names on the
    original PDF (including the proc_2 checkbox trio and the 7-line remarks
    split)."""
    values = {}
    remarks_text = ""

    for field in LOG_COLUMNS:
        if field["id"] == "export_flag":
            continue
        col = field_to_col.get(field["id"])
        if col is None:
            continue
        raw = cell_to_str(ws.cell(row=row_num, column=col).value)
        if not raw:
            continue

        if field["ftype"] == "choice":
            allowed = field["choices"]
            match = next((c for c in allowed if c.lower() == raw.lower()), None)
            if match is None:
                print(f"   ! row {row_num}: '{raw}' is not a valid value for "
                      f"\"{field['label']}\" (expected one of {allowed}) - leaving blank.")
                continue
            raw = match

        if field["id"] in PROC_CHECKBOXES:
            box = PROC_CHECKBOXES[field["id"]].get(raw)
            if box:
                values[box] = CHECKBOX_ON
            continue

        if field["id"] == "remarks":
            remarks_text = raw
            continue

        pdf_field = FIELD_MAP.get(field["id"])
        if pdf_field is None:
            continue  # this item has no real fillable field on the original PDF
        values[pdf_field] = raw

    if remarks_text:
        lines = remarks_text.splitlines() or [remarks_text]
        for line, pdf_field in zip(lines, REMARKS_LINE_FIELDS):
            if line.strip():
                values[pdf_field] = line.strip()

    return values


def stamp_signature(writer):
    """Overlays the Yanda QA Representative signature image onto page 2,
    just above the Signature: line in Part 8."""
    if not SIGNATURE_IMAGE.exists():
        print(f"   ! Signature image not found at {SIGNATURE_IMAGE} - skipping signature stamp. "
              f"Make sure the 'assets' folder is in the same directory as this script.")
        return
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=letter)
    c.drawImage(str(SIGNATURE_IMAGE), SIGNATURE_X, SIGNATURE_Y,
                width=SIGNATURE_W, height=SIGNATURE_H,
                mask="auto", preserveAspectRatio=True)
    c.save()
    buf.seek(0)
    overlay_reader = PdfReader(buf)
    writer.pages[SIGNATURE_PAGE_INDEX].merge_page(overlay_reader.pages[0])


def ensure_default_resources(writer):
    """The original PDF's /AcroForm has no /DR (default resources) entry,
    which crashes pypdf's field-update code, and an empty /DR silently
    breaks --flatten (it can't find a /Helv font to render the flattened
    text with). Populate a real Helvetica font resource instead."""
    acro = writer._root_object["/AcroForm"]
    helv = DictionaryObject({
        NameObject("/Type"): NameObject("/Font"),
        NameObject("/Subtype"): NameObject("/Type1"),
        NameObject("/BaseFont"): NameObject("/Helvetica"),
        NameObject("/Encoding"): NameObject("/WinAnsiEncoding"),
    })
    helv_ref = writer._add_object(helv)
    font_dict = DictionaryObject({NameObject("/Helv"): helv_ref})
    dr = DictionaryObject({NameObject("/Font"): font_dict})
    acro[NameObject("/DR")] = dr
    if "/DA" not in acro:
        acro[NameObject("/DA")] = TextStringObject("/Helv 0 Tf 0 g")


def fill_pdf(template_path, values, out_path, flatten=False, add_signature=True):
    reader = PdfReader(str(template_path))
    writer = PdfWriter()
    writer.append(reader)
    ensure_default_resources(writer)

    for page in writer.pages:
        writer.update_page_form_field_values(page, values, flatten=flatten)
    writer.set_need_appearances_writer(not flatten)

    if add_signature:
        stamp_signature(writer)

    if flatten:
        # update_page_form_field_values(..., flatten=True) only bakes the
        # appearance into the page content - it deliberately leaves the
        # interactive widgets in place (that's documented pypdf behaviour).
        # Actually removing them, plus clearing the AcroForm's field list,
        # is what makes the PDF stop being an editable form.
        writer.remove_annotations(subtypes="/Widget")
        if "/AcroForm" in writer._root_object:
            acro = writer._root_object["/AcroForm"]
            acro[NameObject("/Fields")] = ArrayObject()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "wb") as fh:
        writer.write(fh)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("excel_path", nargs="?", default=str(HERE / "Equipment_Inspection_Tracker.xlsx"),
                     help="Path to the log workbook (default: Transmitter_Inspection_Log.xlsx next to this script)")
    ap.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="Path to the blank fillable PDF template")
    ap.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Folder to write filled PDFs into")
    ap.add_argument("--rows", default="", help="Comma-separated Excel row numbers to export "
                                                 "(overrides the Export to PDF Y/N flag), e.g. --rows 4,7,12")
    ap.add_argument("--all", action="store_true", help="Export every row that has a Tag filled in, "
                                                          "ignoring the Export to PDF Y/N flag")
    ap.add_argument("--merge", action="store_true", help="Also write one combined PDF of everything exported")
    ap.add_argument("--flatten", action="store_true", help="Flatten the filled fields into static page content "
                                                              "(no longer editable/fillable afterward)")
    ap.add_argument("--sheet", default=SHEET_NAME,
                     help=f"Which sheet to read from, e.g. 'Transmitter Log 100' or "
                          f"'Transmitter Log 200' (default: '{SHEET_NAME}')")
    ap.add_argument("--suffix", default="DEV.",
                     help='Text to append to the filename after the Tag, e.g. --suffix "JUG" '
                          'produces "29103-PIT-1021 JUG.pdf". Pass --suffix "" for just "<tag>.pdf". '
                          'Default: "DEV." (matches this script\'s original behavior).')
    ap.add_argument("--no-signature", action="store_true",
                     help="Skip the Yanda QA Representative signature stamp entirely")
    args = ap.parse_args()

    excel_path = Path(args.excel_path)
    template_path = Path(args.template)
    output_dir = Path(args.output_dir)

    if not excel_path.exists():
        sys.exit(f"ERROR: can't find workbook: {excel_path}")
    if not template_path.exists():
        sys.exit(f"ERROR: can't find PDF template: {template_path}")

    explicit_rows = [int(x) for x in args.rows.split(",") if x.strip()] if args.rows else []

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        sys.exit(f"ERROR: workbook has no '{args.sheet}' sheet. "
                  f"Available sheets: {', '.join(wb.sheetnames)}")
    ws = wb[args.sheet]

    field_to_col = load_column_map(ws)
    rows = rows_to_export(ws, field_to_col, explicit_rows, args.all)

    if not rows:
        print("No rows to export. Mark a row's \"Export to PDF (Y/N)\" column as Y, "
              "or pass --rows 4,7,... , or pass --all.")
        return

    print(f"Exporting {len(rows)} row(s): {rows}")
    written = []
    tag_col = field_to_col.get("tag")
    used_names = set()

    for row_num in rows:
        values = build_values_for_row(ws, field_to_col, row_num)
        tag = cell_to_str(ws.cell(row=row_num, column=tag_col).value) if tag_col else ""
        tag_part = sanitize(tag, f"Row{row_num}")
        base_name = f"{tag_part} {args.suffix}" if args.suffix else tag_part
        name = base_name
        n = 2
        while name in used_names:
            name = f"{base_name} ({n})"
            n += 1
        used_names.add(name)
        out_path = output_dir / f"{name}.pdf"
        fill_pdf(template_path, values, out_path, flatten=args.flatten,
                  add_signature=not args.no_signature)
        written.append(out_path)
        print(f"   -> {out_path}")

    if args.merge and written:
        merged_writer = PdfWriter()
        for p in written:
            merged_writer.append(PdfReader(str(p)))
        merged_writer.set_need_appearances_writer(True)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        merged_path = output_dir / f"Combined_Export_{stamp}.pdf"
        with open(merged_path, "wb") as fh:
            merged_writer.write(fh)
        print(f"Combined PDF -> {merged_path}")

    print(f"Done. {len(written)} PDF(s) written to {output_dir}/")


if __name__ == "__main__":
    main()
