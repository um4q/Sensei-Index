#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Turns one or more rows of Valve_Check_Record_Log.xlsx into filled,
ready-to-print PDFs (one PDF per row), using
Pneumatically_Actuated_Valve_Check_Record_TEMPLATE.pdf as the blank form.

BASIC USE (exports every row whose "Export to PDF (Y/N)" column is Y):
    python3 export_valve_to_pdf.py Valve_Check_Record_Log.xlsx

EXPORT SPECIFIC ROWS ONLY (Excel row numbers, overrides the Y/N flag):
    python3 export_valve_to_pdf.py Valve_Check_Record_Log.xlsx --rows 4,7,12

EXPORT EVERY FILLED-IN ROW REGARDLESS OF THE FLAG:
    python3 export_valve_to_pdf.py Valve_Check_Record_Log.xlsx --all

ALSO PRODUCE ONE COMBINED PDF OF EVERYTHING EXPORTED THIS RUN:
    python3 export_valve_to_pdf.py Valve_Check_Record_Log.xlsx --merge

FLATTEN THE OUTPUT (bakes the values into the page, no longer fillable):
    python3 export_valve_to_pdf.py Valve_Check_Record_Log.xlsx --flatten

Requires: pip install openpyxl pypdf
"""
import argparse
import datetime
import re
import sys
from pathlib import Path

import openpyxl
from pypdf import PdfReader, PdfWriter
from pypdf.generic import DictionaryObject, NameObject, TextStringObject, ArrayObject
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import letter
import io

from valve_schema import LOG_COLUMNS
from valve_field_map import (
    FIELD_MAP, CHECKBOX_GROUPS, YES_NO_CHECKBOXES, FV_YES_NA_FIELDS,
    TRAVEL_UNIT_CHECKBOXES, TRAVEL_UNIT_VALUE_FIELDS, COMMENTS_LINE_FIELDS,
    CHECKBOX_ON, CHECKBOX_OFF,
)

# Frozen-aware: when bundled into an .exe by PyInstaller, __file__ points
# inside a temp extraction folder that's wiped on exit - sys.executable's
# folder is the exe's real, persistent location.
HERE = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) \
    else Path(__file__).resolve().parent
DEFAULT_TEMPLATE = HERE / "Pneumatically_Actuated_Valve_Check_Record_TEMPLATE.pdf"
DEFAULT_OUTPUT_DIR = HERE / "output_pdfs"
HEADER_ROW = 3          # combined workbook has a nav-button row above the header
FIRST_DATA_ROW = 4
SHEET_NAME = "Valve Log"

# YANDA QC Representative signature + date stamp - placed on page 1, in the
# QC Representative Sign-Off block (row spans x=38.5-309.7; Signature row
# y=65.4-80.0, label ends x=75.2; Date row y=80.0-94.8, label ends x=57.8).
SIGNATURE_IMAGE = HERE / "assets" / "yanda_qa_signature_transparent.png"
SIGNATURE_PAGE_INDEX = 0  # this form is a single page
SIGNATURE_X = 90
SIGNATURE_Y = 67
SIGNATURE_W = 78
SIGNATURE_H = 78 * (90 / 458)  # preserve the source image's aspect ratio
DATE_STAMP_X = 62
DATE_STAMP_Y = 84.5
DATE_STAMP_FONT_SIZE = 8


def sanitize(text, fallback):
    text = (text or "").strip()
    text = re.sub(r"[^A-Za-z0-9\-_. ]+", "", text).strip()
    return text or fallback


def cell_to_str(value):
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_column_map(ws):
    label_to_col = {}
    dupe_labels = set()
    for col in range(1, ws.max_column + 1):
        label = ws.cell(row=HEADER_ROW, column=col).value
        if label:
            label = str(label).strip()
            if label in label_to_col:
                dupe_labels.add(label)
            label_to_col[label] = col
    if dupe_labels:
        print("WARNING: these column headers appear more than once in row "
              f"{HEADER_ROW} - only the last occurrence of each will be read:")
        for d in sorted(dupe_labels):
            print("   -", d)

    field_to_col = {}
    missing = []
    for field in LOG_COLUMNS:
        col = label_to_col.get(field["label"])
        if col is None:
            missing.append(field["label"])
        else:
            field_to_col[field["id"]] = col
    if missing:
        print("WARNING: these expected columns were not found in the Log sheet header "
              f"(row {HEADER_ROW}) and will be left blank on export:")
        for m in missing:
            print("   -", m)
    return field_to_col


def rows_to_export(ws, field_to_col, explicit_rows, export_all):
    if explicit_rows:
        return sorted(set(explicit_rows))

    export_col = field_to_col.get("export_flag")
    equip_col = field_to_col.get("equip_number")
    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        flag = cell_to_str(ws.cell(row=r, column=export_col).value).upper() if export_col else ""
        has_equip = equip_col and cell_to_str(ws.cell(row=r, column=equip_col).value)
        if export_all:
            if has_equip:
                rows.append(r)
        elif flag == "Y":
            rows.append(r)
    return rows


def build_values_for_row(ws, field_to_col, row_num):
    """Returns a dict of REAL PDF field name -> value."""
    values = {}
    comments_text = ""

    for field in LOG_COLUMNS:
        fid = field["id"]
        if fid == "export_flag":
            continue
        col = field_to_col.get(fid)
        if col is None:
            continue
        raw = cell_to_str(ws.cell(row=row_num, column=col).value)
        if not raw:
            continue

        if field["ftype"] == "choice":
            allowed = field["choices"]
            match = next((c for c in allowed if c.lower() == raw.lower()), None)
            if match is None:
                print(f"   ! row {row_num}: '{raw}' is not a valid value for "
                      f"\"{field['label']}\" (expected one of {allowed}) - leaving blank.")
                continue
            raw = match

        # --- special cases -----------------------------------------------
        if fid in CHECKBOX_GROUPS:
            box = CHECKBOX_GROUPS[fid].get(raw)
            if box:
                values[box] = CHECKBOX_ON
            continue

        if fid in YES_NO_CHECKBOXES:
            if raw.upper() == "Y":
                values[YES_NO_CHECKBOXES[fid]] = CHECKBOX_ON
            continue

        if fid == "travel_unit":
            box = TRAVEL_UNIT_CHECKBOXES.get(raw)
            if box:
                values[box] = CHECKBOX_ON
            continue

        if fid == "travel_value":
            unit_col = field_to_col.get("travel_unit")
            unit_raw = cell_to_str(ws.cell(row=row_num, column=unit_col).value) if unit_col else ""
            target = TRAVEL_UNIT_VALUE_FIELDS.get(unit_raw, TRAVEL_UNIT_VALUE_FIELDS["Inch"])
            values[target] = raw
            continue

        if field["ftype"] == "initial_or_na":
            pair = FV_YES_NA_FIELDS.get(fid)
            if pair:
                if raw.upper() in ("N/A", "NA"):
                    values[pair["na"]] = "N/A"
                else:
                    values[pair["yes"]] = raw
            continue

        if fid == "comments":
            comments_text = raw
            continue

        pdf_field = FIELD_MAP.get(fid)
        if pdf_field is None:
            continue
        values[pdf_field] = raw

    if comments_text:
        lines = comments_text.splitlines() or [comments_text]
        for line, pdf_field in zip(lines, COMMENTS_LINE_FIELDS):
            if line.strip():
                values[pdf_field] = line.strip()

    return values


def ensure_default_resources(writer):
    """The original PDF's /AcroForm has no /DR (default resources) entry,
    which crashes pypdf's field-update code, and an empty /DR silently
    breaks --flatten (it can't find a /Helv font to render the flattened
    text with). Populate a real Helvetica font resource instead."""
    acro = writer._root_object["/AcroForm"]
    helv = DictionaryObject({
        NameObject("/Type"): NameObject("/Font"),
        NameObject("/Subtype"): NameObject("/Type1"),
        NameObject("/BaseFont"): NameObject("/Helvetica"),
        NameObject("/Encoding"): NameObject("/WinAnsiEncoding"),
    })
    helv_ref = writer._add_object(helv)
    font_dict = DictionaryObject({NameObject("/Helv"): helv_ref})
    dr = DictionaryObject({NameObject("/Font"): font_dict})
    acro[NameObject("/DR")] = dr
    if "/DA" not in acro:
        acro[NameObject("/DA")] = TextStringObject("/Helv 0 Tf 0 g")


def widen_choice_options(writer, values):
    """The fv_1..fv_8 Yes/NA dropdowns ship with a fixed preset options list
    (LO / N/A / FS / blank) from whoever last used this template, but in
    practice any inspector's initials should be accepted. Some viewers only
    render a choice field's value correctly if it's actually in that field's
    /Opt list (even though these are 'editable' combo boxes that are
    technically allowed to hold arbitrary text) - so we extend /Opt with
    whatever new value we're about to write, rather than relying on every
    viewer to handle the edit flag the same way."""
    from pypdf.generic import ArrayObject as _Arr, TextStringObject as _Txt
    for page in writer.pages:
        annots = page.get("/Annots")
        if not annots:
            continue
        for a in annots:
            obj = a.get_object()
            name = obj.get("/T")
            if name is None:
                continue
            name = str(name)
            if name not in values or obj.get("/FT") != "/Ch":
                continue
            new_val = values[name]
            opt = obj.get("/Opt")
            if opt is None:
                continue
            current = [str(o) for o in opt]
            if new_val not in current:
                opt.append(_Txt(new_val))


def stamp_signature(writer, sign_date=None):
    """Overlays the YANDA QC Representative signature image (and, unless
    disabled, a date) onto page 1, in the QC Representative Sign-Off block."""
    if not SIGNATURE_IMAGE.exists():
        print(f"   ! Signature image not found at {SIGNATURE_IMAGE} - skipping signature stamp. "
              f"Make sure the 'assets' folder is in the same directory as this script.")
        return
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=letter)
    c.drawImage(str(SIGNATURE_IMAGE), SIGNATURE_X, SIGNATURE_Y,
                width=SIGNATURE_W, height=SIGNATURE_H,
                mask="auto", preserveAspectRatio=True)
    if sign_date:
        c.setFont("Helvetica", DATE_STAMP_FONT_SIZE)
        c.drawString(DATE_STAMP_X, DATE_STAMP_Y, sign_date)
    c.save()
    buf.seek(0)
    overlay_reader = PdfReader(buf)
    writer.pages[SIGNATURE_PAGE_INDEX].merge_page(overlay_reader.pages[0])


def fill_pdf(template_path, values, out_path, flatten=False, sign_date=None, add_signature=True):
    reader = PdfReader(str(template_path))
    writer = PdfWriter()
    writer.append(reader)
    ensure_default_resources(writer)
    widen_choice_options(writer, values)

    for page in writer.pages:
        writer.update_page_form_field_values(page, values, flatten=flatten)
    writer.set_need_appearances_writer(not flatten)

    if add_signature:
        stamp_signature(writer, sign_date=sign_date)

    if flatten:
        writer.remove_annotations(subtypes="/Widget")
        if "/AcroForm" in writer._root_object:
            acro = writer._root_object["/AcroForm"]
            acro[NameObject("/Fields")] = ArrayObject()

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "wb") as fh:
        writer.write(fh)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("excel_path", nargs="?", default=str(HERE / "Equipment_Inspection_Tracker.xlsx"),
                     help="Path to the log workbook (default: Valve_Check_Record_Log.xlsx next to this script)")
    ap.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="Path to the blank fillable PDF template")
    ap.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Folder to write filled PDFs into")
    ap.add_argument("--rows", default="", help="Comma-separated Excel row numbers to export "
                                                 "(overrides the Export to PDF Y/N flag), e.g. --rows 4,7,12")
    ap.add_argument("--all", action="store_true", help="Export every row that has an Equip # filled in, "
                                                          "ignoring the Export to PDF Y/N flag")
    ap.add_argument("--merge", action="store_true", help="Also write one combined PDF of everything exported")
    ap.add_argument("--flatten", action="store_true", help="Flatten the filled fields into static page content "
                                                              "(no longer editable/fillable afterward)")
    ap.add_argument("--sign-date", default="",
                     help="Date to stamp next to the QC Representative signature IMAGE, e.g. "
                          "2026-08-24. Blank by default - the QC Representative's actual sign-off "
                          "date belongs in the 'YANDA QC Representative - Date' column now (a real "
                          "field, filled the normal way from the Excel row), so this flag only "
                          "still matters if you specifically want an extra date drawn next to the "
                          "signature image itself.")
    ap.add_argument("--no-signature", action="store_true",
                     help="Skip the QC Representative signature stamp entirely")
    ap.add_argument("--suffix", default="",
                     help="Text to append to the filename after the Equip #/tag, e.g. --suffix \"DEV.\" "
                          "produces \"29103-PV-1011 DEV..pdf\". Default: no suffix, just \"<tag>.pdf\".")
    ap.add_argument("--sheet", default=SHEET_NAME,
                     help=f"Which sheet to read from, e.g. 'Valve Log 100' or 'Valve Log 200' "
                          f"(default: '{SHEET_NAME}')")
    args = ap.parse_args()

    excel_path = Path(args.excel_path)
    template_path = Path(args.template)
    output_dir = Path(args.output_dir)

    if not excel_path.exists():
        sys.exit(f"ERROR: can't find workbook: {excel_path}")
    if not template_path.exists():
        sys.exit(f"ERROR: can't find PDF template: {template_path}")

    explicit_rows = [int(x) for x in args.rows.split(",") if x.strip()] if args.rows else []

    sign_date = args.sign_date or None

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        sys.exit(f"ERROR: workbook has no '{args.sheet}' sheet. "
                  f"Available sheets: {', '.join(wb.sheetnames)}")
    ws = wb[args.sheet]

    field_to_col = load_column_map(ws)
    rows = rows_to_export(ws, field_to_col, explicit_rows, args.all)

    if not rows:
        print("No rows to export. Mark a row's \"Export to PDF (Y/N)\" column as Y, "
              "or pass --rows 4,7,... , or pass --all.")
        return

    print(f"Exporting {len(rows)} row(s): {rows}")
    written = []
    equip_col = field_to_col.get("equip_number")
    used_names = set()

    for row_num in rows:
        values = build_values_for_row(ws, field_to_col, row_num)
        equip = cell_to_str(ws.cell(row=row_num, column=equip_col).value) if equip_col else ""
        tag_part = sanitize(equip, f"Row{row_num}")
        base_name = f"{tag_part} {args.suffix}" if args.suffix else tag_part
        name = base_name
        n = 2
        while name in used_names:
            name = f"{base_name} ({n})"
            n += 1
        used_names.add(name)
        out_path = output_dir / f"{name}.pdf"
        fill_pdf(template_path, values, out_path, flatten=args.flatten,
                  sign_date=sign_date, add_signature=not args.no_signature)
        written.append(out_path)
        print(f"   -> {out_path}")

    if args.merge and written:
        merged_writer = PdfWriter()
        for p in written:
            merged_writer.append(PdfReader(str(p)))
        merged_writer.set_need_appearances_writer(True)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        merged_path = output_dir / f"Combined_Export_{stamp}.pdf"
        with open(merged_path, "wb") as fh:
            merged_writer.write(fh)
        print(f"Combined PDF -> {merged_path}")

    print(f"Done. {len(written)} PDF(s) written to {output_dir}/")


if __name__ == "__main__":
    main()
