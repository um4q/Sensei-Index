# -*- coding: utf-8 -*-
"""
Data-access layer for the InstINDEX GUI.

Every place in this whole add-on that touches Equipment_Inspection_Tracker.xlsx
or series_registry.json lives in THIS file. gui_app.py never imports openpyxl
directly - it only calls functions defined here. That one rule is what makes
it possible to add a brand-new equipment type later (say, "Gauges") without
touching a single line of gui_app.py: you'd write gauge_schema.py +
gauge_field_map.py + export_gauge_to_pdf.py the same way the existing
Transmitter/Valve trio is built, then add ONE new entry to EQUIPMENT_TYPES
below. Everything else - the menu, the index table, the edit form, the
detail-view PDF - is generic and reads from that registry.

Reused, not duplicated: the column-header lookup, the row-to-PDF-field
translation, and the actual PDF filling all come from the *existing*
export_to_pdf.py / export_valve_to_pdf.py modules (imported below). This file
never re-implements that logic - it just calls it with a specific sheet name
and row number instead of running the whole command-line script.
"""
import contextlib
import datetime
import json
import os
import re
import shutil
import subprocess
import sys
from contextlib import contextmanager
from pathlib import Path

import openpyxl
from pypdf import PdfReader, PdfWriter

import transmitter_schema
import valve_schema
import export_to_pdf
import export_valve_to_pdf
import master_list_reader
# Re-exported here (not duplicated) so Phase 6 and any other caller finds
# these where the v2.1 spec says to look for them - data_access.py - while
# the actual implementation lives in master_list_reader.py, which has no
# dependency on this module (avoids a circular import: this module needs
# master_list_reader for the parser itself, further down).
from master_list_reader import canonical_tag, parse_area_code, classify_kind  # noqa: F401

# When PyInstaller freezes this into a single .exe, __file__ points inside a
# temporary extraction folder (sys._MEIPASS) that's deleted when the app
# closes - using it here would silently lose every edit, setting, and
# signature the moment the exe exits. sys.executable's folder is the actual,
# persistent location of the .exe (or of python.exe when running as a plain
# script, which is why the frozen check comes first).
if getattr(sys, "frozen", False):
    HERE = Path(sys.executable).resolve().parent
else:
    HERE = Path(__file__).resolve().parent

WORKBOOK_PATH = HERE / "Equipment_Inspection_Tracker.xlsx"
CONFIG_PATH = HERE / "series_registry.json"
TEMP_DIR = HERE / "temp_previews"



_wb_cache = {False: None, True: None, "mtime": None}


def _workbook_mtime():
    try:
        return WORKBOOK_PATH.stat().st_mtime
    except OSError:
        return None


def write_json_atomic(path, data):
    """Every JSON file this app writes (series_registry.json,
    app_settings.json, equipment_status.json, and every new sidecar added
    since) goes through this. Two things it guarantees:

    1. Atomicity: the new content is written to a temp file in the SAME
       directory (so it's on the same filesystem - os.replace() is only
       atomic within one filesystem) and only swapped into place with
       os.replace() once the write fully succeeds. A crash, power loss, or
       killed process mid-save leaves either the old file untouched or the
       complete new one - never a truncated/half-written file that bricks
       the next startup.
    2. A permissions problem always gives the same clear explanation
       instead of a raw traceback. In practice this only ever fires one
       way: the app's own folder needs administrator rights to write to -
       almost always because it's sitting inside Program Files or another
       protected location, which normal saves inside the app hit
       constantly.
    """
    path = Path(path)
    tmp_path = path.with_name(f".{path.name}.tmp{os.getpid()}")
    try:
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, indent=2)
        os.replace(tmp_path, path)
    except PermissionError:
        with contextlib.suppress(OSError):
            tmp_path.unlink()
        raise PermissionError(
            f"Can't write to '{path.name}'. This app's folder needs administrator "
            "rights to write to - almost always because it's inside Program Files "
            "or another protected location. Move the whole InstINDEX folder "
            "somewhere any account can write to without admin, like "
            "C:\\Users\\Public\\InstINDEX, your Desktop, or Documents."
        ) from None
    except Exception:
        with contextlib.suppress(OSError):
            tmp_path.unlink()
        raise


# Populated by read_json_with_recovery() whenever it has to reset a
# corrupted sidecar to defaults. gui_app checks this once at startup and
# shows each entry as a toast - so "the file got reset" is never silent,
# but it also never blocks the app from opening the way a crash would.
STARTUP_WARNINGS = []


def read_json_with_recovery(path, default):
    """Loads JSON from path, tolerating both a missing file (returns
    default, no warning - normal for a sidecar that hasn't been created
    yet) and a corrupted one (renames the bad file to
    '<name>.corrupt.<timestamp>' - preserved for hand recovery, never
    silently deleted - queues a STARTUP_WARNINGS entry, and returns
    default). default may be a plain value or a zero-arg callable (use a
    callable when the default is a mutable container, so every caller
    doesn't share one dict/list instance)."""
    path = Path(path)

    def _default():
        return default() if callable(default) else default

    if not path.exists():
        return _default()
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except (json.JSONDecodeError, OSError, UnicodeDecodeError):
        stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        corrupt_path = path.with_name(f"{path.name}.corrupt.{stamp}")
        try:
            path.rename(corrupt_path)
            STARTUP_WARNINGS.append(
                f"'{path.name}' was corrupted and has been reset to defaults. "
                f"The unreadable file was saved as '{corrupt_path.name}' next to it."
            )
        except OSError:
            STARTUP_WARNINGS.append(
                f"'{path.name}' was corrupted and could not be read - using defaults."
            )
        return _default()


def _get_cached_workbook(data_only=False):
    """Returns a shared, live openpyxl Workbook - reloaded from disk only
    when the file's mtime has changed since the last load. Callers must NOT
    call .close() on what this returns; the cache owns its lifecycle.
    Mutating callers (save_row, add_series, delete_rows, remove_series,
    ...) must call _save_workbook_and_refresh_cache(wb) afterward instead
    of saving directly, so the cache and the on-disk file never disagree."""
    current_mtime = _workbook_mtime()
    if current_mtime != _wb_cache["mtime"]:
        # File is new to us, or changed since we last looked (our own save,
        # or someone editing it directly in Excel) - drop both cached views
        # so they're reloaded fresh, lazily, below.
        _wb_cache[False] = None
        _wb_cache[True] = None
        _wb_cache["mtime"] = current_mtime
    if _wb_cache[data_only] is None:
        _wb_cache[data_only] = openpyxl.load_workbook(WORKBOOK_PATH, data_only=data_only)
    return _wb_cache[data_only]


def _save_workbook_and_refresh_cache(wb):
    """Saves a workbook obtained from _get_cached_workbook(data_only=False)
    and updates the cache to match, so the very next read (e.g. the reload
    that follows almost every save in the GUI) doesn't pay for a disk
    round-trip it doesn't need. The data_only=True view is dropped rather
    than refreshed - openpyxl can't recompute formulas itself, so the only
    honest thing to do with a 'last calculated value' cache after an edit is
    to let it reload fresh the next time something actually needs it."""
    _save_workbook(wb)
    _wb_cache[False] = wb
    _wb_cache[True] = None
    _wb_cache["mtime"] = _workbook_mtime()


def invalidate_workbook_cache():
    """Forces the next read to reload from disk. Not needed in normal use
    (saves already refresh the cache themselves) - exposed mainly so 'Open
    Excel File' + manual edits followed by a Refresh in the app can't ever
    show stale data if mtime granularity ever lies on some filesystem."""
    _wb_cache[False] = None
    _wb_cache[True] = None
    _wb_cache["mtime"] = None


@contextmanager
def _mutating_workbook():
    """Yields the shared cached workbook for a block of in-place edits,
    saving it and refreshing the cache on success. If anything inside the
    block raises, the in-memory workbook is thrown away (never saved) and
    the cache is invalidated - otherwise a half-finished edit (say, the
    Transmitter sheet got copied for a new series but the Valve sheet copy
    then failed) would sit in memory looking like valid, current data on
    every subsequent read, when what's actually on disk never changed.

    Phase 16.1: also the single choke point for automatic backups - a
    snapshot of the CURRENT on-disk file is taken here, before this
    block's write, whenever the last one is older than
    backup_interval_minutes (or there isn't one yet). Never blocks the
    actual mutation: a backup failure (disk full, permissions) is warned
    to console and swallowed."""
    _backup_workbook_if_due()
    wb = _get_cached_workbook(data_only=False)
    try:
        yield wb
    except Exception:
        invalidate_workbook_cache()
        raise
    else:
        _save_workbook_and_refresh_cache(wb)


# ---------------------------------------------------------------------------
# Equipment type registry - the single place that defines what a
# "Transmitter" or a "Valve" IS, as far as the GUI is concerned.
# ---------------------------------------------------------------------------
EQUIPMENT_TYPES = {
    "transmitter": {
        "label": "Transmitter",
        "schema": transmitter_schema,
        "export_module": export_to_pdf,
        "key_field": "tag",
        "summary_fields": ["tag", "system_number", "transmitter_type"],
        "summary_labels": ["Tag", "System #", "Type"],
        "group_fields": ["system_number"],
        "group_labels": ["System"],
        "date_fields": ["te1_caldate", "te2_caldate", "te3_caldate", "yanda_qa_date", "client_date"],
        "date_labels": ["TE1 Cal. Date", "TE2 Cal. Date", "TE3 Cal. Date", "QA Rep Date", "Client Rep Date"],
    },
    "valve": {
        "label": "Valve",
        "schema": valve_schema,
        "export_module": export_valve_to_pdf,
        "key_field": "equip_number",
        "summary_fields": ["equip_number", "system", "valve_type"],
        "summary_labels": ["Equip #", "System", "Type"],
        "group_fields": ["system"],
        "group_labels": ["System"],
        "date_fields": ["equip_caldate", "qc_date"],
        "date_labels": ["TE Cal. Date", "QC Rep Date"],
    },
}

ASSETS_DIR = HERE / "assets"
SETTINGS_PATH = HERE / "app_settings.json"
STATUS_PATH = HERE / "equipment_status.json"


# ---------------------------------------------------------------------------
# series_registry.json
# ---------------------------------------------------------------------------
def load_config():
    return read_json_with_recovery(CONFIG_PATH, lambda: {"series": []})


def save_config(cfg):
    write_json_atomic(CONFIG_PATH, cfg)


def list_series():
    """e.g. [100, 200]"""
    return sorted(s["number"] for s in load_config()["series"])


def get_sheet_name(series_number, equip_key):
    for s in load_config()["series"]:
        if s["number"] == series_number:
            key = f"{equip_key}_sheet"
            if key not in s:
                raise KeyError(f"Series {series_number} has no '{equip_key}' sheet registered.")
            return s[key]
    raise KeyError(f"Series {series_number} is not in {CONFIG_PATH.name}")


# ---------------------------------------------------------------------------
# Series display names
#
# Series have only ever had a number. This adds an optional friendly label
# next to it (stored as "name" on the series' entry in series_registry.json)
# without touching the workbook at all - so it's instant, and there's no
# sheet-renaming risk to anything Excel-side (named ranges, formulas
# referencing the sheet by name, etc).
# ---------------------------------------------------------------------------
def get_series_name(series_number):
    for s in load_config()["series"]:
        if s["number"] == series_number:
            return s.get("name", "")
    return ""


def set_series_name(series_number, name):
    cfg = load_config()
    for s in cfg["series"]:
        if s["number"] == series_number:
            name = (name or "").strip()
            if name:
                s["name"] = name
            else:
                s.pop("name", None)
            save_config(cfg)
            return
    raise KeyError(f"Series {series_number} is not in {CONFIG_PATH.name}")


def series_display_label(series_number):
    """The custom name if one's been set, or just the bare series number
    as a fallback. Deliberately never says the word 'Series' here - this
    is what actually shows up throughout the sidebar, dashboard, and index
    page titles day-to-day, so a name (or, failing that, just the number)
    reads a lot cleaner than 'Series 100' everywhere."""
    name = get_series_name(series_number)
    return name if name else str(series_number)


# ---------------------------------------------------------------------------
# Removing a whole series
#
# Un-registers it from the app and archives (renames + hides) its
# Transmitter/Valve Log sheets rather than deleting them outright. These are
# equipment inspection/QA records - a single misclick permanently destroying
# a series' entire history is a much worse failure mode than a hidden,
# oddly-named leftover sheet sitting in the workbook that a human can find
# and restore by hand in Excel (Home tab, right-click a sheet tab, Unhide).
# ---------------------------------------------------------------------------
def _archived_sheet_name(original_name, wb):
    """A sheet name under Excel's hard 31-character limit that still says
    what it is and when it was removed, however long the original name."""
    stamp = datetime.date.today().strftime("%y%m%d")
    prefix = "DEL "
    budget = 31 - len(prefix) - len(stamp) - 1  # -1 for the space before the stamp
    candidate = f"{prefix}{original_name[:budget]} {stamp}"
    base, n = candidate, 2
    while candidate in wb.sheetnames:
        suffix = f"-{n}"
        candidate = base[:31 - len(suffix)] + suffix
        n += 1
    return candidate


def remove_series(series_number):
    """Raises KeyError if the series isn't registered. Returns nothing on
    success - callers should refresh whatever series list they're showing."""
    cfg = load_config()
    entry = next((s for s in cfg["series"] if s["number"] == series_number), None)
    if entry is None:
        raise KeyError(f"Series {series_number} is not in {CONFIG_PATH.name}")

    with _mutating_workbook() as wb:
        for equip_key in EQUIPMENT_TYPES:
            sheet_name = entry.get(f"{equip_key}_sheet")
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.title = _archived_sheet_name(sheet_name, wb)
                ws.sheet_state = "hidden"

    cfg["series"] = [s for s in cfg["series"] if s["number"] != series_number]
    save_config(cfg)

    prefix = f"{series_number}|"
    store = _load_status_store()
    filtered = {k: v for k, v in store.items() if not k.startswith(prefix)}
    if len(filtered) != len(store):
        _save_status_store(filtered)


# ---------------------------------------------------------------------------
# Reading rows for the Index view
# ---------------------------------------------------------------------------
def _looks_like_excel_serial(text):
    """True for a bare integer roughly in the range Excel's date serials
    cover for realistic years (~1954-2064). Used only to decide whether a
    date field's raw text is worth trying to reformat for display - never
    to guess at non-date fields."""
    t = text.strip()
    if not t or not t.lstrip("-").isdigit():
        return False
    n = int(t)
    return 19000 <= n <= 60000


def format_date_for_display(raw_text):
    """Best-effort cosmetic fix for date fields whose cell was never given
    a proper Excel date format - some rows in this workbook have calibration
    / sign-off dates stored as a bare serial number (e.g. "46234") that
    Excel itself would also show as "46234" until the cell's format is
    fixed. Converts anything that looks like one of those into a plain
    "YYYY-MM-DD" string for display. Purely cosmetic: the underlying cell
    is untouched unless that field is actually edited afterward, at which
    point whatever gets typed (or Mass Edit Dates sets) is saved as clean
    text and the formatting problem is gone for that cell for good."""
    if not raw_text:
        return raw_text
    if _looks_like_excel_serial(raw_text):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(int(raw_text)).strftime("%Y-%m-%d")
        except Exception:
            return raw_text
    return raw_text


def read_index_rows(series_number, equip_key):
    """[{'row': 5, 'tag': '...', 'system_number': '...', ...}, ...] - one
    dict per row that has its key field (Tag / Equip #) filled in. Includes
    both the equipment type's summary_fields AND date_fields - the Index
    table's grid shows and mass-edits both."""
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    sheet_name = get_sheet_name(series_number, equip_key)

    wb = _get_cached_workbook(data_only=False)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Workbook has no sheet named '{sheet_name}'")
    ws = wb[sheet_name]
    field_to_col = export_mod.load_column_map(ws)
    key_col = field_to_col.get(etype["key_field"])
    date_fields = set(etype.get("date_fields", []))
    display_fields = list(etype["summary_fields"]) + list(date_fields)

    rows = []
    for r in range(export_mod.FIRST_DATA_ROW, ws.max_row + 1):
        key_val = ws.cell(row=r, column=key_col).value if key_col else None
        if key_val in (None, ""):
            continue
        entry = {"row": r}
        for fid in display_fields:
            col = field_to_col.get(fid)
            raw = ws.cell(row=r, column=col).value if col else ""
            text = export_mod.cell_to_str(raw)
            entry[fid] = format_date_for_display(text) if fid in date_fields else text
        rows.append(entry)
    return rows


def read_full_row(series_number, equip_key, row_num):
    """{field_id: value_as_string} for every schema field on one row - used
    to pre-fill the Edit form."""
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    schema = etype["schema"]
    sheet_name = get_sheet_name(series_number, equip_key)
    date_fields = set(etype.get("date_fields", []))

    values = {f["id"]: "" for f in schema.LOG_COLUMNS}

    wb = _get_cached_workbook(data_only=False)
    ws = wb[sheet_name]
    field_to_col = export_mod.load_column_map(ws)
    for fid, col in field_to_col.items():
        text = export_mod.cell_to_str(ws.cell(row=row_num, column=col).value)
        values[fid] = format_date_for_display(text) if fid in date_fields else text
    return values


def find_first_blank_row(series_number, equip_key):
    """First data row whose key field (Tag / Equip #) is empty - used for
    'Add New'."""
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    sheet_name = get_sheet_name(series_number, equip_key)

    wb = _get_cached_workbook(data_only=False)
    ws = wb[sheet_name]
    field_to_col = export_mod.load_column_map(ws)
    key_col = field_to_col.get(etype["key_field"])
    r = export_mod.FIRST_DATA_ROW
    while key_col and ws.cell(row=r, column=key_col).value not in (None, ""):
        r += 1
    return r


def save_row(series_number, equip_key, row_num, values, source="edit_dialog"):
    """Writes values (field_id -> string) into one row of the matching Log
    sheet and saves the workbook. Only touches columns the schema knows
    about - every other cell (styling, the yellow example-row formatting,
    other columns) is left exactly as it was.

    Phase 15 choke point: logs ONE activity_log.jsonl event for this row
    (only if at least one field actually changed value - a no-op save
    logs nothing)."""
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    sheet_name = get_sheet_name(series_number, equip_key)
    key_field = etype["key_field"]

    changed = {}
    with _mutating_workbook() as wb:
        ws = wb[sheet_name]
        field_to_col = export_mod.load_column_map(ws)
        for fid, val in values.items():
            col = field_to_col.get(fid)
            if col is None:
                continue
            old_text = export_mod.cell_to_str(ws.cell(row=row_num, column=col).value)
            new_text = val if val != "" else ""
            if old_text != new_text:
                changed[fid] = {"old": old_text, "new": new_text}
            ws.cell(row=row_num, column=col).value = (val if val != "" else None)

    if changed:
        key_value = values.get(key_field) or changed.get(key_field, {}).get("new", "")
        log_activity("edit", series=series_number, equip_key=equip_key, row=row_num,
                      key_value=key_value, fields=changed, source=source)


def save_fields_bulk(series_number, equip_key, updates, source="bulk"):
    """updates: an iterable of (row_num, field_id, value) triples - writes
    all of them within a SINGLE workbook load/save cycle, regardless of how
    many rows or fields are touched. This is the efficient path behind
    every "mass edit" action in the Index grid (paste across a column,
    fill-down, mass-clear, Mass Edit Dates) - the alternative, calling
    save_row() once per cell, would mean one full disk write per cell,
    which gets slow fast once a paste spans dozens of rows.

    Phase 15 choke point: logs ONE activity_log.jsonl event PER ROW
    touched (not one for the whole batch, not one per field) - all
    sharing one identical timestamp, per 15's documented convention."""
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    sheet_name = get_sheet_name(series_number, equip_key)
    key_field = etype["key_field"]

    changed_by_row = {}
    with _mutating_workbook() as wb:
        ws = wb[sheet_name]
        field_to_col = export_mod.load_column_map(ws)
        for row_num, fid, val in updates:
            col = field_to_col.get(fid)
            if col is None:
                continue
            old_text = export_mod.cell_to_str(ws.cell(row=row_num, column=col).value)
            new_text = val if val != "" else ""
            if old_text != new_text:
                changed_by_row.setdefault(row_num, {})[fid] = {"old": old_text, "new": new_text}
            ws.cell(row=row_num, column=col).value = (val if val != "" else None)
        # Key values for the log, read from the SAME open sheet/cache -
        # never a second workbook round-trip just for logging.
        key_col = field_to_col.get(key_field)
        key_values_by_row = {
            row_num: export_mod.cell_to_str(ws.cell(row=row_num, column=key_col).value)
            for row_num in changed_by_row
        } if key_col else {}

    if changed_by_row:
        shared_ts = datetime.datetime.now().isoformat(timespec="seconds")
        for row_num, fields in changed_by_row.items():
            key_value = key_values_by_row.get(row_num) or fields.get(key_field, {}).get("new", "")
            log_activity("edit", series=series_number, equip_key=equip_key, row=row_num,
                          key_value=key_value, fields=fields, source=source, ts=shared_ts)


# ---------------------------------------------------------------------------
# Removing rows
#
# This clears every schema-mapped cell on each given row rather than
# physically deleting the worksheet row (ws.delete_rows). A real delete
# would shift every row below it up by one - which sounds fine until you
# remember data-validation dropdown ranges and conditional-formatting rules
# (like the Export-flag green highlight) are anchored to specific row
# numbers, and openpyxl's support for re-anchoring those on a shift is
# known to be incomplete. Clearing in place touches only the cells the
# schema already owns, can't disturb anything else on the sheet, and the
# row becomes indistinguishable from an unused one - find_first_blank_row()
# will naturally hand it back out the next time something is added, so
# nothing is wasted.
# ---------------------------------------------------------------------------
def delete_rows(series_number, equip_key, row_nums, source="edit_dialog"):
    """Clears the given row number(s) in one series+type's Log sheet.
    Returns the list of key-field values (Tag / Equip #) that were cleared,
    so their Installed/Submitted status entries can be dropped too.

    Phase 15 choke point: logs one 'delete' event per row actually
    cleared (had a key value) - the field-by-field diff isn't logged
    here (deleting can blank a hundred-plus fields at once; the ACTION
    itself, and which tag it was, is what per-row History needs)."""
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    sheet_name = get_sheet_name(series_number, equip_key)
    key_field = etype["key_field"]

    cleared_keys = []
    cleared_rows = []  # (row_num, key_val) pairs, for logging below
    with _mutating_workbook() as wb:
        ws = wb[sheet_name]
        field_to_col = export_mod.load_column_map(ws)
        key_col = field_to_col.get(key_field)

        for row_num in row_nums:
            if key_col:
                key_val = ws.cell(row=row_num, column=key_col).value
                if key_val not in (None, ""):
                    cleared_keys.append(str(key_val))
                    cleared_rows.append((row_num, str(key_val)))
            for col in field_to_col.values():
                ws.cell(row=row_num, column=col).value = None

    if cleared_rows:
        shared_ts = datetime.datetime.now().isoformat(timespec="seconds")
        for row_num, key_val in cleared_rows:
            log_activity("delete", series=series_number, equip_key=equip_key, row=row_num,
                          key_value=key_val, source=source, ts=shared_ts)

    if cleared_keys:
        store = _load_status_store()
        changed = False
        for key_val in cleared_keys:
            if store.pop(_status_key(series_number, equip_key, key_val), None) is not None:
                changed = True
        if changed:
            _save_status_store(store)

    return cleared_keys


class WorkbookLockedError(Exception):
    """Raised in place of the raw PermissionError when the workbook
    appears to be open in Excel (or another program holding an exclusive
    lock) at save time (16.2) - callers show a specific 'close it, then
    Retry' dialog for this, instead of the generic Program-Files/admin-
    rights PermissionError message."""


def _excel_lock_file_path():
    # Excel's own convention: '~$<filename>' next to the real file,
    # created the moment the workbook is opened and removed on a clean
    # close. A stale one (Excel/the OS crashed) is possible but rare - a
    # false positive here just means one extra Retry click, never data
    # loss - so it's used as a hint, not sole proof.
    return WORKBOOK_PATH.with_name(f"~${WORKBOOK_PATH.name}")


def is_workbook_locked():
    """Best-effort, never raises - a check that itself fails is treated
    as 'not locked' (the real answer comes from the save attempt either
    way). The '~$...' lockfile is the portable signal; on Windows, an
    exclusive-rename probe backs it up for the case Excel didn't leave
    one (or it was already cleaned up) but still holds the file open."""
    try:
        if _excel_lock_file_path().exists():
            return True
    except OSError:
        pass
    if os.name == "nt" and WORKBOOK_PATH.exists():
        try:
            # Renaming a file to its own name fails on Windows if any
            # process has it open with a sharing lock that blocks
            # renames - which Excel's own open-for-edit lock does.
            os.rename(WORKBOOK_PATH, WORKBOOK_PATH)
        except OSError:
            return True
    return False


def _save_workbook(wb):
    try:
        wb.save(WORKBOOK_PATH)
    except PermissionError:
        if is_workbook_locked():
            raise WorkbookLockedError(
                "The workbook is open in Excel. Close it there, then click Retry."
            ) from None
        raise PermissionError(
            f"Can't save '{WORKBOOK_PATH.name}'. One of two things is almost "
            "always the cause: it's open in Excel (or another program) right "
            "now - close it there and try again - or this app's folder needs "
            "administrator rights to write to, which usually means it's "
            "sitting inside Program Files or another protected location. If "
            "it's the second one, move the whole InstINDEX folder somewhere "
            "any account can write to without admin, like "
            "C:\\Users\\Public\\InstINDEX, your Desktop, or Documents."
        ) from None


# ---------------------------------------------------------------------------
# "View Details" - fill one row into a PDF in a temp folder, then open it
# ---------------------------------------------------------------------------
def generate_preview_pdf(series_number, equip_key, row_num):
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    sheet_name = get_sheet_name(series_number, equip_key)

    TEMP_DIR.mkdir(exist_ok=True)
    for old in TEMP_DIR.glob("*.pdf"):          # keep the temp folder from growing forever
        try:
            old.unlink()
        except OSError:
            pass  # a previous preview is probably still open in a PDF viewer - fine, skip it

    wb = _get_cached_workbook(data_only=True)
    ws = wb[sheet_name]
    field_to_col = export_mod.load_column_map(ws)
    values = export_mod.build_values_for_row(ws, field_to_col, row_num)
    key_col = field_to_col.get(etype["key_field"])
    key_val = export_mod.cell_to_str(ws.cell(row=row_num, column=key_col).value) if key_col else ""

    safe_name = export_mod.sanitize(key_val, f"row{row_num}")
    out_path = TEMP_DIR / f"{safe_name} PREVIEW.pdf"

    if equip_key == "valve":
        export_mod.fill_pdf(export_mod.DEFAULT_TEMPLATE, values, out_path,
                             flatten=False, add_signature=True)
    else:
        export_mod.fill_pdf(export_mod.DEFAULT_TEMPLATE, values, out_path, flatten=False)

    return out_path


def open_file(path):
    """Opens a file with whatever the OS considers its default app for it."""
    path = str(path)
    if sys.platform.startswith("win"):
        os.startfile(path)  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.run(["open", path])
    else:
        subprocess.run(["xdg-open", path])


# ---------------------------------------------------------------------------
# Settings -> Add a new series
# ---------------------------------------------------------------------------
def add_series(new_number, source_number=None):
    """Duplicates an existing series' Transmitter + Valve Log sheets into a
    brand-new, empty series, and registers it in series_registry.json.

    Best-effort: openpyxl's copy_worksheet() copies cell values, styles,
    column widths, and merged cells - but it silently DROPS data-validation
    dropdowns and conditional formatting. This function re-attaches both by
    hand afterward, then wipes the data rows so the new sheet starts empty.

    If this ever produces a sheet where a dropdown doesn't work, the
    reliable fallback is to duplicate the sheet manually in Excel (right-click
    the tab -> Move or Copy -> check "Create a copy") and then call
    register_existing_series() instead.
    """
    cfg = load_config()
    existing_numbers = [s["number"] for s in cfg["series"]]
    if new_number in existing_numbers:
        raise ValueError(f"Series {new_number} already exists.")
    if source_number is None:
        source_number = existing_numbers[0]

    new_entry = {"number": new_number}
    with _mutating_workbook() as wb:
        for equip_key, etype in EQUIPMENT_TYPES.items():
            src_name = get_sheet_name(source_number, equip_key)
            new_name = f"{etype['label']} Log {new_number}"
            if new_name in wb.sheetnames:
                raise ValueError(f"Sheet '{new_name}' already exists in the workbook.")

            src_ws = wb[src_name]
            new_ws = wb.copy_worksheet(src_ws)
            new_ws.title = new_name

            for dv in src_ws.data_validations.dataValidation:
                new_ws.add_data_validation(dv)
            for cf in src_ws.conditional_formatting:
                for rule in cf.rules:
                    new_ws.conditional_formatting.add(str(cf.sqref), rule)

            export_mod = etype["export_module"]
            for r in range(export_mod.FIRST_DATA_ROW, new_ws.max_row + 1):
                for c in range(1, new_ws.max_column + 1):
                    new_ws.cell(row=r, column=c).value = None

            new_entry[f"{equip_key}_sheet"] = new_name

    cfg["series"].append(new_entry)
    save_config(cfg)
    return new_entry


def register_existing_series(new_number, transmitter_sheet, valve_sheet):
    """For the manual fallback: you already created 'Transmitter Log 300' /
    'Valve Log 300' yourself in Excel - this just teaches the app about
    them, with no sheet copying at all."""
    wb = _get_cached_workbook(data_only=False)
    for name in (transmitter_sheet, valve_sheet):
        if name not in wb.sheetnames:
            raise KeyError(f"Workbook has no sheet named '{name}'")

    cfg = load_config()
    if any(s["number"] == new_number for s in cfg["series"]):
        raise ValueError(f"Series {new_number} already exists.")
    cfg["series"].append({
        "number": new_number,
        "transmitter_sheet": transmitter_sheet,
        "valve_sheet": valve_sheet,
    })
    save_config(cfg)


# ---------------------------------------------------------------------------
# Instructions + live counts (for the Instructions dialog and the Main Menu /
# Series Menu summary numbers)
# ---------------------------------------------------------------------------
def read_instructions():
    """Column B of the workbook's 'Instructions' sheet, one string per
    non-empty row, in order. Returns [] if the sheet doesn't exist."""
    wb = _get_cached_workbook(data_only=False)
    if "Instructions" not in wb.sheetnames:
        return []
    ws = wb["Instructions"]
    lines = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val:
            lines.append(str(val))
    return lines


def count_rows(series_number, equip_key):
    """How many rows in this series+type have their key field filled in."""
    return len(read_index_rows(series_number, equip_key))


def series_type_summary(series_number, equip_key, group_field):
    """(count, {group_value: count}) in one pass over read_index_rows() -
    used by the sidebar, which used to call count_rows() and
    distinct_group_values() back to back (two separate scans of the same
    rows) for every series x equipment-type combo it draws."""
    rows = read_index_rows(series_number, equip_key)
    counts = {}
    for r in rows:
        val = str(r.get(group_field) or "").strip()
        if val:
            counts[val] = counts.get(val, 0) + 1
    return len(rows), dict(sorted(counts.items()))


def count_all_by_type():
    """{'transmitter': <total across every series>, 'valve': <...>}"""
    totals = {k: 0 for k in EQUIPMENT_TYPES}
    for series_number in list_series():
        for equip_key in EQUIPMENT_TYPES:
            try:
                totals[equip_key] += count_rows(series_number, equip_key)
            except KeyError:
                pass  # this series doesn't register that equipment type
    return totals


# ---------------------------------------------------------------------------
# Per-series Installed/Submitted/Accepted breakdown - powers the Dashboard's
# "BY SERIES" cards (one card per series, each showing how many Transmitters
# and Valves are logged AND how many of each are Installed/Submitted/
# Accepted - the plain totals in count_all_by_type() above only ever showed
# a single grand total per equipment type, with no per-series or
# per-status view at all).
# ---------------------------------------------------------------------------
def series_full_summary(series_number):
    """{'transmitter': {'total': N, 'installed': n, 'submitted': n, 'accepted': n},
         'valve': {...}} for ONE series. An equipment type is simply absent
    from the result if this series has no sheet registered for it."""
    result = {}
    for equip_key in EQUIPMENT_TYPES:
        try:
            rows = read_index_rows_with_status(series_number, equip_key)
        except KeyError:
            continue
        result[equip_key] = {
            "total": len(rows),
            "installed": sum(1 for r in rows if r.get("installed")),
            "submitted": sum(1 for r in rows if r.get("submitted")),
            "accepted": sum(1 for r in rows if r.get("accepted")),
        }
    return result


# ---------------------------------------------------------------------------
# Populating Wizard - a single resumable draft.
#
# Deliberately just ONE slot (not a named/multi-draft library): the wizard
# is meant for "I'm about to bulk-enter a batch of similar equipment, got
# interrupted, and want to pick the batch back up later" - not for
# maintaining a permanent library of in-progress templates. Saving a new
# draft overwrites whatever was there before, and the wizard tells you
# that plainly before it lets you do so.
# ---------------------------------------------------------------------------
DRAFTS_PATH = HERE / "wizard_draft.json"


def load_wizard_draft():
    return read_json_with_recovery(DRAFTS_PATH, None)


def save_wizard_draft(data):
    write_json_atomic(DRAFTS_PATH, data)


def clear_wizard_draft():
    try:
        DRAFTS_PATH.unlink()
    except FileNotFoundError:
        pass


# ---------------------------------------------------------------------------
# Phase 10 - Master List import engine.
#
# master_list.json is a normalized SNAPSHOT of the last-imported
# Instrumentation_Master_List.xlsx - read-only reference data. The app
# never writes back into the user's copy of that workbook; every write
# here only ever touches master_list.json next to the app, same as any
# other sidecar.
# ---------------------------------------------------------------------------
MASTER_LIST_PATH = HERE / "master_list.json"


def load_master_list():
    """The last-imported snapshot, or None if nothing has ever been
    imported (or the snapshot was corrupt and got reset - see
    read_json_with_recovery, which already queues the STARTUP_WARNINGS
    entry for that case). Callers should treat None exactly like "no
    master list" - Coverage's empty state, search's empty master-list
    group, etc."""
    return read_json_with_recovery(MASTER_LIST_PATH, None)


def save_master_list(data):
    write_json_atomic(MASTER_LIST_PATH, data)


def preview_master_list(path):
    """Read-only parse of a candidate master list file for the import
    dialog's step 1 - writes nothing. Returns (items_by_sheet, summaries),
    exactly master_list_reader.read_master_list_file()'s return. A thin
    pass-through kept here (rather than gui_app.py importing
    master_list_reader directly) so gui_app.py keeps calling data_access
    for every data operation, per this module's own design contract."""
    return master_list_reader.read_master_list_file(path)


def guess_master_list_series_mapping(sheet_names):
    """{sheet_name: series_number or None} pre-guessed against this app's
    OWN registered series numbers (finding 1.8) - the import dialog's
    step 2 starting point, always still a user-editable combo per sheet."""
    return master_list_reader.guess_series_mapping(sheet_names, list_series())


def import_master_list(path, sheet_series_map):
    """Parses the given .xlsx (master_list_reader.read_master_list_file),
    stamps every item with its sheet's mapped series number from
    sheet_series_map ({sheet_name: series_number_or_None} - None for a
    sheet the user left unmapped, finding 1.8), and atomically writes
    master_list.json - fully REPLACING whatever snapshot was there before
    (re-importing is idempotent: it never accumulates stale rows from an
    earlier version of the source file). Also remembers the path in
    app_settings.json (master_list_path) for one-click re-import.

    Returns (snapshot, summaries): snapshot is exactly what got saved;
    summaries is the parser's per-sheet list (rows found, transmitter/
    valve/out_of_scope counts, problems) for the import dialog's step-1
    review table."""
    items_by_sheet, summaries = master_list_reader.read_master_list_file(path)

    items = []
    for sheet_name, sheet_items in items_by_sheet.items():
        mapped_series = sheet_series_map.get(sheet_name)
        for item in sheet_items:
            item = dict(item)
            item["mapped_series"] = mapped_series
            items.append(item)

    source_path = Path(path)
    snapshot = {
        "imported_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "source_file": source_path.name,
        "source_mtime": source_path.stat().st_mtime,
        "sheet_series_map": dict(sheet_series_map),
        "items": items,
    }
    save_master_list(snapshot)
    set_setting("master_list_path", str(source_path))
    invalidate_search_index()
    log_activity("master_list_import", source="master_list",
                  note=f"Imported {len(items)} row(s) from {source_path.name}")
    return snapshot, summaries


def master_list_items(kind=None, area=None, series=None):
    """Filtered view over the last imported master list. [] if nothing has
    been imported yet. `series` filters on mapped_series (a sheet's
    confirmed target series number, not the raw sheet name)."""
    snapshot = load_master_list()
    if not snapshot:
        return []
    items = snapshot.get("items", [])
    if kind is not None:
        items = [it for it in items if it.get("kind") == kind]
    if area is not None:
        items = [it for it in items if it.get("area") == area]
    if series is not None:
        items = [it for it in items if it.get("mapped_series") == series]
    return items


def master_list_needs_reimport():
    """True when app_settings.json remembers a master_list_path, that file
    still exists on disk, and its mtime is newer than what's recorded in
    the last-saved snapshot - powers the Coverage page's 'Master list has
    changed on disk - re-import?' banner (10.4/11.2). False (never
    crashes) if nothing's ever been imported, or the remembered path is
    gone."""
    remembered_path = get_setting("master_list_path")
    if not remembered_path:
        return False
    source_path = Path(remembered_path)
    if not source_path.exists():
        return False
    snapshot = load_master_list()
    if not snapshot:
        return False
    try:
        return source_path.stat().st_mtime > snapshot.get("source_mtime", 0)
    except OSError:
        return False


# ---------------------------------------------------------------------------
# Phase 13.3 - sticky per-user IndexPage view state: last sort column/
# order, active filter chip, and (for the columns a user can actually
# resize - see gui_app.py) column widths. Keyed by "<series>:<equip_key>"
# so each series+type page remembers its own state independently.
# ---------------------------------------------------------------------------
UI_STATE_PATH = HERE / "ui_state.json"


def _ui_state_page_key(series_number, equip_key):
    return f"{series_number}:{equip_key}"


def load_ui_state():
    return read_json_with_recovery(UI_STATE_PATH, dict)


def save_ui_state(data):
    write_json_atomic(UI_STATE_PATH, data)


def get_page_view_state(series_number, equip_key):
    """{} if this page has never saved any view state - callers apply
    whatever keys are present and fall back to their own defaults for
    whatever's missing, never crash on a partial or absent entry."""
    state = load_ui_state()
    return state.get(_ui_state_page_key(series_number, equip_key), {})


def set_page_view_state(series_number, equip_key, **fields):
    """Merges fields into this page's saved state (doesn't require the
    caller to pass every key every time) and writes the whole store back
    atomically."""
    state = load_ui_state()
    key = _ui_state_page_key(series_number, equip_key)
    page_state = dict(state.get(key, {}))
    page_state.update(fields)
    state[key] = page_state
    save_ui_state(state)


# ---------------------------------------------------------------------------
# Phase 15 - Activity log & per-row history.
#
# Append-only JSONL, one JSON object per line: {"ts", "action", "series",
# "equip_key", "row", "key_value", "fields": {field_id: {"old","new"}},
# "source", "note"}. Hooked at the CHOKE POINTS below (save_row,
# save_fields_bulk, set_status/bulk_set_status, delete_rows,
# create_rows_from_master_items, import_master_list) - never per-widget -
# so every caller (grid edit, EditDialog, Populating Wizard, datasheet
# import, create-from-master, undo/redo) gets logged automatically without
# gui_app.py having to remember to call anything.
#
# Convention (documented per 15's own acceptance criterion): a bulk action
# touching N rows produces N events, one per row, all sharing one ts and
# source - never one combined event for the whole batch, and never one
# event per individual field either.
#
# Writing history NEVER raises - a disk-full/permissions failure here is
# warned to console and swallowed; history must never block a save.
# ---------------------------------------------------------------------------
ACTIVITY_LOG_PATH = HERE / "activity_log.jsonl"
ACTIVITY_LOG_ROTATE_BYTES = 5 * 1024 * 1024  # ~5MB
ACTIVITY_LOG_FIELD_TRUNCATE = 200


def _truncate_for_log(value):
    text = "" if value is None else str(value)
    return text if len(text) <= ACTIVITY_LOG_FIELD_TRUNCATE else text[:ACTIVITY_LOG_FIELD_TRUNCATE] + "…"


def _rotate_activity_log_if_needed():
    try:
        if ACTIVITY_LOG_PATH.exists() and ACTIVITY_LOG_PATH.stat().st_size > ACTIVITY_LOG_ROTATE_BYTES:
            rotated = ACTIVITY_LOG_PATH.with_name("activity_log.1.jsonl")
            if rotated.exists():
                rotated.unlink()  # keep exactly one prior generation
            ACTIVITY_LOG_PATH.rename(rotated)
    except OSError as exc:
        print(f"WARNING: activity_log.jsonl rotation failed (continuing without rotating): {exc}")


def log_activity(action, series=None, equip_key=None, row=None, key_value=None,
                  fields=None, source="app", note=None, ts=None):
    """Appends one event. fields: {field_id: {'old': ..., 'new': ...}} -
    each value truncated to ACTIVITY_LOG_FIELD_TRUNCATE chars. ts lets a
    caller share one identical timestamp across a batch of per-row events
    (save_fields_bulk does this) - omit it to stamp 'now'."""
    entry = {
        "ts": ts or datetime.datetime.now().isoformat(timespec="seconds"),
        "action": action, "series": series, "equip_key": equip_key,
        "row": row, "key_value": key_value,
        "fields": {
            fid: {"old": _truncate_for_log(diff.get("old")), "new": _truncate_for_log(diff.get("new"))}
            for fid, diff in (fields or {}).items()
        },
        "source": source, "note": note,
    }
    try:
        _rotate_activity_log_if_needed()
        with open(ACTIVITY_LOG_PATH, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(entry) + "\n")
    except OSError as exc:
        print(f"WARNING: couldn't write to activity_log.jsonl (continuing without logging): {exc}")


def read_activity_log(limit=None, series=None, equip_key=None, row=None):
    """Newest-first. Tolerates a corrupt/truncated line (skips it, never
    crashes) and a missing file (returns []) - deleting or hand-editing
    the log must never break the app (15's own acceptance criterion)."""
    if not ACTIVITY_LOG_PATH.exists():
        return []
    entries = []
    try:
        with open(ACTIVITY_LOG_PATH, "r", encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if series is not None and entry.get("series") != series:
                    continue
                if equip_key is not None and entry.get("equip_key") != equip_key:
                    continue
                if row is not None and entry.get("row") != row:
                    continue
                entries.append(entry)
    except OSError:
        return []
    entries.reverse()
    return entries[:limit] if limit is not None else entries


def read_row_history(series_number, equip_key, row_num, key_value=None, limit=None):
    """Every event for one row - RowDetailDialog's 'History' tab. Matches
    by row NUMBER (stable across a Tag/Equip # rename) primarily; status
    events don't carry a row number (set_status is keyed by key_value
    alone, not a row), so entries matching key_value are merged in too,
    deduplicated by identity. A renamed row's PRE-rename status events
    (logged under the old key_value) won't appear here - the same known
    limitation equipment_status.json itself has without
    rename_status_key()."""
    by_row = read_activity_log(series=series_number, equip_key=equip_key, row=row_num)
    by_key = [e for e in read_activity_log(series=series_number, equip_key=equip_key)
              if e.get("row") is None and e.get("key_value") == key_value] if key_value else []
    merged = by_row + by_key
    merged.sort(key=lambda e: e.get("ts") or "", reverse=True)
    return merged[:limit] if limit is not None else merged


# ---------------------------------------------------------------------------
# Phase 11 - Coverage / reconciliation.
#
# Answers, per area and per equipment kind, which master-list instruments
# have a tracker row and which don't - and the reverse. Matching is by
# canonical_tag ONLY, across every registered series of the matching kind
# (a master item isn't required to belong to its sheet's mapped_series to
# be found - the tag might already be tracked under a different series
# number by mistake, and that's worth surfacing too, not silently missed).
# ---------------------------------------------------------------------------
def _normalize_for_compare(value):
    """case/space-insensitive comparison key for the Model mismatch flag
    (11.1) - not a display value."""
    return re.sub(r"\s+", "", str(value or "")).casefold()


def _tracker_refs_by_tag(equip_key):
    """{canonical_tag: {'series', 'equip_key', 'row', 'key_value',
    'installed', 'submitted', 'accepted'}} across every registered series
    for one equipment kind. First occurrence wins on an accidental
    duplicate canonical tag (same convention as the master-list parser)."""
    etype = EQUIPMENT_TYPES[equip_key]
    key_field = etype["key_field"]
    refs = {}
    for series_number in list_series():
        try:
            rows = read_index_rows_with_status(series_number, equip_key)
        except KeyError:
            continue
        for entry in rows:
            tag = canonical_tag(entry.get(key_field, ""))
            if not tag or tag in refs:
                continue
            refs[tag] = {
                "series": series_number, "equip_key": equip_key, "row": entry["row"],
                "key_value": entry.get(key_field, ""), "installed": bool(entry.get("installed")),
                "submitted": bool(entry.get("submitted")), "accepted": bool(entry.get("accepted")),
            }
    return refs


def reconcile_master_list():
    """The Coverage page's core computation. Returns None if nothing has
    ever been imported (mirrors load_master_list()'s own convention -
    callers show Coverage's empty state in that case, never a crash).
    Otherwise:

    {'matched': [{'master': item, 'tracker': ref, 'flags': [str, ...],
                  'progress': {...compute_progress() shape...}}, ...],
     'missing': [item, ...],          # in-scope master items, no tracker row anywhere
     'orphans': [ref, ...],           # tracker rows whose tag isn't in the master list
     'out_of_scope_count': int,
     'out_of_scope_items': [item, ...]}   # shown collapsed (count only), expandable

    ref (a tracker-side row reference): {'series', 'equip_key', 'row',
    'key_value', 'installed', 'submitted', 'accepted'}. flags are
    machine-readable ids ('installed_mismatch', 'model_mismatch') - the UI
    supplies wording. 'missing' items have no tracker row at all, so no
    progress dict - callers show a flat "0% - not started" for those.

    matched + missing always == every in-scope (transmitter/valve) master
    item; orphans are counted separately, same as 11's own acceptance
    criteria."""
    snapshot = load_master_list()
    if not snapshot:
        return None

    tracker_refs = {equip_key: _tracker_refs_by_tag(equip_key) for equip_key in EQUIPMENT_TYPES}
    matched_tags = {equip_key: set() for equip_key in EQUIPMENT_TYPES}

    matched, missing, out_of_scope_items = [], [], []

    for item in snapshot.get("items", []):
        kind = item.get("kind")
        if kind not in EQUIPMENT_TYPES:
            out_of_scope_items.append(item)
            continue

        ref = tracker_refs[kind].get(item["tag"])
        if ref is None:
            missing.append(item)
            continue

        matched_tags[kind].add(item["tag"])
        flags = []
        if bool(item.get("ml_installed")) != ref["installed"]:
            flags.append("installed_mismatch")

        tracker_values = read_full_row(ref["series"], kind, ref["row"])
        model_field = "model" if kind == "transmitter" else "valve_model"
        tracker_model = tracker_values.get(model_field, "")
        master_model = item.get("model", "")
        if (master_model and tracker_model
                and _normalize_for_compare(master_model) != _normalize_for_compare(tracker_model)):
            flags.append("model_mismatch")

        # Phase 12.2: Coverage shows the same per-row % everywhere else in
        # the app does - reuses the tracker_values already read above for
        # the model-mismatch check, no extra workbook read.
        progress = compute_progress(kind, tracker_values, ref)

        matched.append({"master": item, "tracker": ref, "flags": flags, "progress": progress})

    orphans = [
        ref for equip_key, refs_by_tag in tracker_refs.items()
        for tag, ref in refs_by_tag.items() if tag not in matched_tags[equip_key]
    ]

    return {
        "matched": matched,
        "missing": missing,
        "orphans": orphans,
        "out_of_scope_count": len(out_of_scope_items),
        "out_of_scope_items": out_of_scope_items,
    }


# ---------------------------------------------------------------------------
# 11.3 - Create-from-master: turn a "Missing" master item into a real
# tracker row. Only maps to fields that actually exist on that equipment
# kind's schema - Valve has no single generic "service"/"manufacturer"
# field the way Transmitter does, so those two are silently omitted for
# valves (the confirm dialog lists exactly what's set, so nothing is
# hidden from the user - there's just nowhere on the valve form for a
# generic Service/Manufacturer value to go).
# ---------------------------------------------------------------------------
def master_item_to_row_values(item, equip_key):
    """{field_id: value} to pre-fill a brand-new tracker row from one
    master-list item - values.py's blank-means-unset convention applies
    (an empty string here is the same as not setting the field at all)."""
    qa_name = get_setting("default_qa_rep_name")
    if equip_key == "valve":
        values = {
            "equip_number": item.get("tag", ""),
            "pid_number": item.get("pid", ""),
            "line_number": item.get("line_number", ""),
            "valve_model": item.get("model", ""),
        }
        if qa_name:
            values["qc_rep_name"] = qa_name
    else:
        values = {
            "tag": item.get("tag", ""),
            "service": item.get("service", ""),
            "pid_number": item.get("pid", ""),
            "line_number": item.get("line_number", ""),
            "make": item.get("manufacturer", ""),
            "model": item.get("model", ""),
        }
        if qa_name:
            values["yanda_qa_name"] = qa_name
    return {k: v for k, v in values.items() if v}


def create_rows_from_master_items(items_with_kind):
    """items_with_kind: [(equip_key, master_item), ...]. Creates one new
    tracker row per entry - find_first_blank_row + the same cell-write
    save_row() does - in a SINGLE workbook save regardless of how many
    series/sheets are touched (save_fields_bulk's efficiency reasoning
    applies here too). This is the same commit path as any other new row:
    there's no separate "pending" staging layer in this app to route
    through instead.

    Raises ValueError before writing anything if any item has no valid
    target series registered - a batch is all-or-nothing, never partially
    applied.

    Returns [{'series', 'equip_key', 'row', 'tag', 'values'}, ...] in the
    same order, for the caller to build ONE undo entry covering the whole
    batch."""
    registered = set(list_series())
    for equip_key, item in items_with_kind:
        if item.get("mapped_series") not in registered:
            raise ValueError(
                f"'{item.get('tag')}' has no valid target series - map its sheet to "
                "a series first (Master List import, step 2).")

    created = []
    with _mutating_workbook() as wb:
        for equip_key, item in items_with_kind:
            series_number = item["mapped_series"]
            sheet_name = get_sheet_name(series_number, equip_key)
            row_num = find_first_blank_row(series_number, equip_key)
            values = master_item_to_row_values(item, equip_key)

            etype = EQUIPMENT_TYPES[equip_key]
            export_mod = etype["export_module"]
            ws = wb[sheet_name]
            field_to_col = export_mod.load_column_map(ws)
            for fid, val in values.items():
                col = field_to_col.get(fid)
                if col is not None:
                    ws.cell(row=row_num, column=col).value = val

            created.append({
                "series": series_number, "equip_key": equip_key, "row": row_num,
                "tag": item.get("tag", ""), "values": values,
            })

    if created:
        shared_ts = datetime.datetime.now().isoformat(timespec="seconds")
        for c in created:
            fields = {fid: {"old": "", "new": val} for fid, val in c["values"].items()}
            log_activity("edit", series=c["series"], equip_key=c["equip_key"], row=c["row"],
                          key_value=c["tag"], fields=fields, source="master_list", ts=shared_ts)
    return created


# ---------------------------------------------------------------------------
# App-wide settings: theme, default export preferences, which signature
# image is "active". Lives in its own small JSON file - nothing to do with
# the workbook at all.
# ---------------------------------------------------------------------------
DEFAULT_SETTINGS = {
    "theme": "bootstrap-light",
    "active_signature": "yanda_qa_signature_transparent.png",
    "default_flatten": False,
    "default_include_signature": True,
    # Sensei Index 2.1 additions:
    "default_qa_rep_name": "",           # pre-fills the QA rep name on create-from-master (Coverage)
    "master_list_path": "",              # remembered path for Master List re-import (Coverage)
    "backup_interval_minutes": 30,       # automatic workbook backup cadence
    "backup_keep": 20,                   # how many workbook snapshots to retain
}


def load_settings():
    data = read_json_with_recovery(SETTINGS_PATH, dict)
    merged = dict(DEFAULT_SETTINGS)
    merged.update(data)
    return merged


def save_settings(settings):
    write_json_atomic(SETTINGS_PATH, settings)


def get_setting(key):
    return load_settings().get(key, DEFAULT_SETTINGS.get(key))


def set_setting(key, value):
    settings = load_settings()
    settings[key] = value
    save_settings(settings)


# ---------------------------------------------------------------------------
# Signature image management (the assets/ folder)
# ---------------------------------------------------------------------------
SIGNATURE_EXTENSIONS = {".png", ".jpg", ".jpeg"}


def list_signatures():
    ASSETS_DIR.mkdir(exist_ok=True)
    return sorted(p.name for p in ASSETS_DIR.iterdir()
                  if p.suffix.lower() in SIGNATURE_EXTENSIONS)


def add_signature(source_path, display_name):
    """Copies an existing PNG/JPG from anywhere on disk into assets/, named
    after display_name. Returns the new filename."""
    source_path = Path(source_path)
    if not source_path.exists():
        raise FileNotFoundError(f"Can't find {source_path}")
    ext = source_path.suffix.lower()
    if ext not in SIGNATURE_EXTENSIONS:
        raise ValueError("Signature file must be a .png, .jpg, or .jpeg image.")

    # Validate with the exact same reader stamp_signature() uses later, so
    # passing validation here is a real guarantee it'll work at export time
    # too - not just a superficial extension check.
    try:
        from reportlab.lib.utils import ImageReader
        ImageReader(str(source_path))
    except Exception as exc:
        raise ValueError(f"That file doesn't look like a valid image ({exc}).") from exc

    safe_name = re.sub(r"[^A-Za-z0-9_\- ]+", "", display_name).strip() or "signature"
    ASSETS_DIR.mkdir(exist_ok=True)
    filename = f"{safe_name}{ext}"
    dest = ASSETS_DIR / filename
    n = 2
    while dest.exists():
        filename = f"{safe_name} ({n}){ext}"
        dest = ASSETS_DIR / filename
        n += 1

    shutil.copy(str(source_path), str(dest))
    if not get_setting("active_signature") or get_active_signature_path() is None:
        set_setting("active_signature", filename)  # first one added becomes active automatically
    return filename


def delete_signature(filename):
    path = ASSETS_DIR / filename
    if path.exists():
        path.unlink()
    if get_setting("active_signature") == filename:
        remaining = list_signatures()
        set_setting("active_signature", remaining[0] if remaining else "")


def get_active_signature_path():
    active = get_setting("active_signature")
    if not active:
        return None
    path = ASSETS_DIR / active
    return path if path.exists() else None


def set_active_signature(filename):
    set_setting("active_signature", filename)


def _apply_active_signature(export_mod):
    """Points the export module's SIGNATURE_IMAGE constant at whichever file
    is currently active in Settings. export_to_pdf.py / export_valve_to_pdf.py
    both read SIGNATURE_IMAGE as a plain module-level global at call time
    (not a baked-in default argument), so reassigning it here from outside
    genuinely changes what the next fill_pdf() call stamps - no need to
    touch either export script for this part."""
    active_path = get_active_signature_path()
    if active_path is not None:
        export_mod.SIGNATURE_IMAGE = active_path


# ---------------------------------------------------------------------------
# Installed / Submitted / Accepted status tracking.
#
# Deliberately kept OUTSIDE the Excel file, in its own JSON store keyed by
# series + equipment type + the row's key-field value (Tag / Equip #), so
# this never requires inserting new columns into your real, in-production
# Log sheets (which would also mean widening every data-validation and
# conditional-formatting range on every existing and future series sheet -
# a much bigger, riskier change for what's meant to be a lightweight
# tracking flag). The trade-off: this status lives in equipment_status.json
# next to the workbook, not as a column you'd see opening the raw .xlsx.
#
# "accepted" was added after installed/submitted, and "export" after that -
# which means real stored entries from earlier updates only have whichever
# keys existed when they were written. Every read of a status dict below
# goes through a DEFAULT_STATUS merge rather than trusting the stored dict
# to already have every key - a plain `[...]` lookup on an old entry would
# KeyError the very first time a pre-existing row's status was read after
# one of these changes.
#
# "export" is the row-level Export checkbox in the Index table: ticking it
# queues that row for the NEXT export run regardless of the Excel sheet's
# own "Export to PDF (Y/N)" column. It's intentionally stored here (next to
# installed/submitted/accepted) rather than written into the workbook, for
# the same reason those are: no new column, no widened data-validation /
# conditional-formatting ranges on every Log sheet.
# ---------------------------------------------------------------------------
DEFAULT_STATUS = {"installed": False, "submitted": False, "accepted": False, "export": False}


def _load_status_store():
    return read_json_with_recovery(STATUS_PATH, dict)


def _save_status_store(store):
    write_json_atomic(STATUS_PATH, store)


def _status_key(series_number, equip_key, key_value):
    return f"{series_number}|{equip_key}|{key_value}"


def _normalized_status(raw):
    """DEFAULT_STATUS, overlaid with whatever fields a stored entry
    actually has - so an entry saved before 'accepted' existed still comes
    back with accepted=False instead of a missing key."""
    merged = dict(DEFAULT_STATUS)
    if raw:
        merged.update(raw)
    return merged


def get_status(series_number, equip_key, key_value):
    store = _load_status_store()
    return _normalized_status(store.get(_status_key(series_number, equip_key, key_value)))


def set_status(series_number, equip_key, key_value, source="app", **fields):
    """Phase 15 choke point: logs one 'status' event, field-diffed the
    same way save_row does (only the fields that actually changed)."""
    store = _load_status_store()
    k = _status_key(series_number, equip_key, key_value)
    before = _normalized_status(store.get(k))
    current = dict(before)
    current.update(fields)
    store[k] = current
    _save_status_store(store)

    changed = {fid: {"old": before[fid], "new": current[fid]} for fid in fields if before.get(fid) != current.get(fid)}
    if changed:
        log_activity("status", series=series_number, equip_key=equip_key,
                      key_value=key_value, fields=changed, source=source)


def bulk_set_status(keys, source="bulk", **fields):
    """keys: an iterable of (series_number, equip_key, key_value) tuples -
    e.g. every row currently selected or visible in a filtered Index
    window. Loads and saves the status file exactly once regardless of how
    many keys are passed, so selecting a hundred rows and toggling a status
    on all of them is still a single read + single write.

    Phase 15 choke point: logs one 'status' event per key actually
    changed, all sharing one timestamp (same convention as
    save_fields_bulk)."""
    store = _load_status_store()
    log_entries = []  # (series, equip_key, key_value, changed_fields)
    for series_number, equip_key, key_value in keys:
        k = _status_key(series_number, equip_key, key_value)
        before = _normalized_status(store.get(k))
        current = dict(before)
        current.update(fields)
        store[k] = current
        changed = {fid: {"old": before[fid], "new": current[fid]}
                   for fid in fields if before.get(fid) != current.get(fid)}
        if changed:
            log_entries.append((series_number, equip_key, key_value, changed))
    _save_status_store(store)

    if log_entries:
        shared_ts = datetime.datetime.now().isoformat(timespec="seconds")
        for series_number, equip_key, key_value, changed in log_entries:
            log_activity("status", series=series_number, equip_key=equip_key,
                          key_value=key_value, fields=changed, source=source, ts=shared_ts)


def rename_status_key(series_number, equip_key, old_key, new_key):
    """Migrates an Installed/Submitted/Accepted/Export status entry from
    old_key to new_key. Called whenever the Tag/Equip # itself is edited
    in place (inline in the Index grid) - without this, the row's status
    would appear to silently reset to unchecked the moment its identifying
    text changed, since the status store is keyed by that exact text."""
    if not old_key or not new_key or old_key == new_key:
        return
    store = _load_status_store()
    old_full = _status_key(series_number, equip_key, old_key)
    new_full = _status_key(series_number, equip_key, new_key)
    if old_full in store:
        store[new_full] = store.pop(old_full)
        _save_status_store(store)


# ---------------------------------------------------------------------------
# Phase 12 - Unified progress model.
#
# One definition of "how done is this instrument," used identically by
# IndexPage, DashboardPage, and Coverage - so the app has a single number
# instead of four disconnected checkboxes (Installed / Submitted / Accepted
# / "did anyone fill out the inspection"). Pure function, no workbook or
# sidecar I/O: callers pass in a read_full_row()-shaped values dict and a
# get_status()-shaped status dict, both of which they already have.
#
# Milestones are equal-weighted by default. PROGRESS_WEIGHTS is a plain
# dict (not baked into the function) specifically so a future "Accepted
# should count for more than Serial captured" ask from the client is a
# one-dict change, not a rewrite (see spec assumption 12.4).
# ---------------------------------------------------------------------------
PROGRESS_MILESTONES = [
    ("created", "Created"),
    ("serial_captured", "Serial captured"),
    ("installed", "Installed"),
    ("inspection_complete", "Inspection complete"),
    ("submitted", "Submitted"),
    ("accepted", "Accepted"),
]

PROGRESS_WEIGHTS = {
    "transmitter": {mid: 1 for mid, _label in PROGRESS_MILESTONES},
    "valve": {mid: 1 for mid, _label in PROGRESS_MILESTONES},
}

# Every valve "component" the check record has a Model/Serial pair for.
# "valve" itself is always required (it's what the row IS); the rest only
# count toward "serial captured" if that component's model was actually
# specified - a row with no positioner shouldn't be penalized for a blank
# positioner serial.
_VALVE_COMPONENTS = ["valve", "actuator", "positioner", "solenoid", "position_limit"]
_TRANSMITTER_VI_FIELDS = [f"vi_{i}" for i in range(1, 13)]         # PART 3, transmitter_schema.py
_TRANSMITTER_PROC_FIELDS = [f"proc_{i}" for i in range(2, 15)]     # PART 4 items 2-14 (excludes proc_1_vdc)
_VALVE_FV_FIELDS = [f"fv_{i}" for i in range(1, 9)]                # Functional Verification, valve_schema.py


def _progress_blank(value):
    return not str(value or "").strip()


def _serial_captured(equip_key, values):
    if equip_key != "valve":
        return not _progress_blank(values.get("serial_number"))
    for component in _VALVE_COMPONENTS:
        model_blank = component != "valve" and _progress_blank(values.get(f"{component}_model"))
        if model_blank:
            continue  # this component isn't present on this valve at all
        if _progress_blank(values.get(f"{component}_serial")):
            return False
    return True


def _inspection_complete(equip_key, values):
    fields = _VALVE_FV_FIELDS if equip_key == "valve" else (_TRANSMITTER_VI_FIELDS + _TRANSMITTER_PROC_FIELDS)
    return all(not _progress_blank(values.get(f)) for f in fields)


def compute_progress(equip_key, values, status):
    """values: {field_id: str}, shaped like read_full_row()'s return.
    status: {'installed': bool, 'submitted': bool, 'accepted': bool, ...},
    shaped like get_status()'s return (missing keys are treated as False,
    so a plain {} also works).

    Returns {'percent': int 0-100,
             'milestones': [{'id', 'label', 'done'}, ...] in fixed order,
             'next': id of the first not-done milestone, or None if every
                     milestone is done}."""
    values = values or {}
    status = status or {}

    done_by_id = {
        "created": True,  # this function is only ever called for a row that exists
        "serial_captured": _serial_captured(equip_key, values),
        "installed": bool(status.get("installed")),
        "inspection_complete": _inspection_complete(equip_key, values),
        "submitted": bool(status.get("submitted")),
        "accepted": bool(status.get("accepted")),
    }

    weights = PROGRESS_WEIGHTS.get(equip_key, PROGRESS_WEIGHTS["transmitter"])
    total_weight = sum(weights.values()) or 1
    earned_weight = sum(weights[mid] for mid, _label in PROGRESS_MILESTONES if done_by_id[mid])
    percent = round(100 * earned_weight / total_weight)

    milestones = [{"id": mid, "label": label, "done": done_by_id[mid]}
                  for mid, label in PROGRESS_MILESTONES]
    next_id = next((mid for mid, _label in PROGRESS_MILESTONES if not done_by_id[mid]), None)

    return {"percent": percent, "milestones": milestones, "next": next_id}


# ---------------------------------------------------------------------------
# Status-enriched + filterable row reading, for the Index window and the
# System/Type drill-down browser. Both build on read_index_rows() - neither
# re-reads the workbook independently.
# ---------------------------------------------------------------------------
def read_index_rows_with_status(series_number, equip_key):
    etype = EQUIPMENT_TYPES[equip_key]
    rows = read_index_rows(series_number, equip_key)
    store = _load_status_store()
    for entry in rows:
        key_val = entry.get(etype["key_field"], "")
        status = _normalized_status(store.get(_status_key(series_number, equip_key, key_val)))
        entry["installed"] = status["installed"]
        entry["submitted"] = status["submitted"]
        entry["accepted"] = status["accepted"]
        entry["export"] = status["export"]
    return rows


def read_index_rows_filtered(series_number, equip_key, filters=None):
    """filters: {field_id: value} - e.g. {'system_number': 'STEAM'}. Applied
    as an exact (string) match, on top of the status-enriched row list."""
    rows = read_index_rows_with_status(series_number, equip_key)
    if filters:
        for fid, val in filters.items():
            rows = [r for r in rows if str(r.get(fid) or "") == str(val)]
    return rows


def read_index_rows_with_progress(series_number, equip_key, filters=None):
    """read_index_rows_with_status()'s rows, each with a 'progress' key
    added ({'percent', 'milestones', 'next'} from compute_progress()).
    Reads every schema field ONCE per row, reusing a single header-derived
    field_to_col map for the whole sheet, rather than calling
    read_full_row() per row (which would re-scan the header every time) -
    the Index page's Progress column (12.2) needs this for every visible
    row, so the batch path matters. filters: same {field_id: value} exact-
    match convention as read_index_rows_filtered() - applied AFTER
    progress is computed, so a filtered view's percentages are identical
    to the unfiltered one for the same rows."""
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    schema = etype["schema"]
    sheet_name = get_sheet_name(series_number, equip_key)

    wb = _get_cached_workbook(data_only=False)
    ws = wb[sheet_name]
    field_to_col = export_mod.load_column_map(ws)

    rows = read_index_rows_with_status(series_number, equip_key)
    for entry in rows:
        row_num = entry["row"]
        values = {}
        for field in schema.LOG_COLUMNS:
            col = field_to_col.get(field["id"])
            values[field["id"]] = export_mod.cell_to_str(ws.cell(row=row_num, column=col).value) \
                if col else ""
        entry["progress"] = compute_progress(equip_key, values, entry)

    if filters:
        for fid, val in filters.items():
            rows = [r for r in rows if str(r.get(fid) or "") == str(val)]
    return rows


def series_progress_summary(series_number):
    """{'transmitter': {'avg_percent', 'at_0', 'partial', 'at_100'}, 'valve':
    {...}} for ONE series - additive to series_full_summary()'s Installed/
    Submitted/Accepted counts, same 'absent if this series has no sheet for
    that type' convention. at_0/partial/at_100 always sum to that type's
    total row count."""
    result = {}
    for equip_key in EQUIPMENT_TYPES:
        try:
            rows = read_index_rows_with_progress(series_number, equip_key)
        except KeyError:
            continue
        if not rows:
            result[equip_key] = {"avg_percent": 0, "at_0": 0, "partial": 0, "at_100": 0}
            continue
        percents = [r["progress"]["percent"] for r in rows]
        result[equip_key] = {
            "avg_percent": round(sum(percents) / len(percents)),
            "at_0": sum(1 for p in percents if p == 0),
            "partial": sum(1 for p in percents if 0 < p < 100),
            "at_100": sum(1 for p in percents if p == 100),
        }
    return result


def distinct_group_values(series_number, equip_key, group_field, filters=None):
    """{value: count} for every distinct non-blank value of group_field,
    among rows matching any filters already chosen at a shallower level of
    the drill-down browser."""
    rows = read_index_rows(series_number, equip_key)
    if filters:
        for fid, val in filters.items():
            rows = [r for r in rows if str(r.get(fid) or "") == str(val)]
    counts = {}
    for r in rows:
        val = str(r.get(group_field) or "").strip()
        if val:
            counts[val] = counts.get(val, 0) + 1
    return dict(sorted(counts.items()))


# ---------------------------------------------------------------------------
# Batch export (the new Export menu) - mirrors what export_to_pdf.py /
# export_valve_to_pdf.py's own main() does, but driven by explicit GUI
# parameters instead of argv, and able to target a subfolder of output_pdfs/.
# ---------------------------------------------------------------------------
def run_export(series_number, equip_key, mode, suffix, flatten, include_signature,
                subfolder=None, merge=False, filters=None, clear_after_selected=True,
                include_date_in_filename=False):
    """mode:
        'selected' - only rows whose Export checkbox is checked in the app
                     (the DEFAULT mode - see ExportDialog). Nothing to do
                     with the Excel sheet at all; this is purely the
                     app-side checkbox state from equipment_status.json.
        'flagged'  - only rows with the Excel "Export to PDF (Y/N)" column
                     set to Y (the sheet's own status, unchanged from the
                     original workbook-driven behavior).
        'all'      - every row with the key field filled in.
    filters, if given (e.g. {'system_number': 'STEAM'}), narrows any of
    these down further to rows matching it - the same filter dict the
    Index page's System drill-down already uses, so "export everything
    currently shown while filtered by system" reuses the exact same
    matching logic as what's on screen.

    clear_after_selected: when mode == 'selected' and this is True (the
    default), every row actually written un-checks its own Export box
    afterward - the queue empties as it's used, like a real "ready to
    send" tray rather than a sticky flag someone has to remember to clear.

    Sign-off dates (both equipment types) come from real schema fields now
    - qc_date for valves, yanda_qa_date/client_date for transmitters - so
    they're just ordinary values from the row like anything else, edited
    ahead of time in the Index grid or with Mass Edit Dates. Valve exports
    used to also silently stamp today's date next to the signature image
    as a separate canvas overlay; that's gone now that qc_date is a real,
    visible field occupying that exact spot on the form - stamping a
    second date there would just draw over the first.

    include_date_in_filename: appends today's date to every output
    filename, e.g. "29103-PIT-1021 DEV. 2026-08-27.pdf".

    Returns the list of Paths written (the merged PDF, if any, is last)."""
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    sheet_name = get_sheet_name(series_number, equip_key)
    _apply_active_signature(export_mod)

    out_dir = HERE / "output_pdfs"
    if subfolder:
        safe_sub = re.sub(r'[<>:"/\\|?*]', "_", subfolder).strip()
        if safe_sub:
            out_dir = out_dir / safe_sub
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = _get_cached_workbook(data_only=True)
    ws = wb[sheet_name]
    field_to_col = export_mod.load_column_map(ws)

    if mode == "selected":
        checked_rows = [r["row"] for r in read_index_rows_filtered(series_number, equip_key, filters)
                         if r.get("export")]
        rows = export_mod.rows_to_export(ws, field_to_col, checked_rows, False)
    else:
        rows = export_mod.rows_to_export(ws, field_to_col, [], mode == "all")
        if filters:
            allowed_rows = {r["row"] for r in read_index_rows_filtered(series_number, equip_key, filters)}
            rows = [r for r in rows if r in allowed_rows]

    if not rows:
        return []

    key_col = field_to_col.get(etype["key_field"])
    filename_date = f" {datetime.date.today().isoformat()}" if include_date_in_filename else ""
    written = []
    used_names = set()
    for row_num in rows:
        values = export_mod.build_values_for_row(ws, field_to_col, row_num)
        key_val = export_mod.cell_to_str(ws.cell(row=row_num, column=key_col).value) if key_col else ""
        tag_part = export_mod.sanitize(key_val, f"Row{row_num}")
        base_name = f"{tag_part} {suffix}" if suffix else tag_part
        base_name = f"{base_name}{filename_date}"
        name = base_name
        n = 2
        while name in used_names:
            name = f"{base_name} ({n})"
            n += 1
        used_names.add(name)
        out_path = out_dir / f"{name}.pdf"

        if equip_key == "valve":
            export_mod.fill_pdf(export_mod.DEFAULT_TEMPLATE, values, out_path, flatten=flatten,
                                 add_signature=include_signature)
        else:
            export_mod.fill_pdf(export_mod.DEFAULT_TEMPLATE, values, out_path, flatten=flatten,
                                 add_signature=include_signature)
        written.append(out_path)

    if merge and written:
        merged_writer = PdfWriter()
        for p in written:
            merged_writer.append(PdfReader(str(p)))
        merged_writer.set_need_appearances_writer(True)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        merged_path = out_dir / f"Combined_Export_{stamp}.pdf"
        with open(merged_path, "wb") as fh:
            merged_writer.write(fh)
        written.append(merged_path)

    if mode == "selected" and clear_after_selected:
        key_col = field_to_col.get(etype["key_field"])
        keys = []
        for row_num in rows:
            key_val = export_mod.cell_to_str(ws.cell(row=row_num, column=key_col).value) if key_col else ""
            if key_val:
                keys.append((series_number, equip_key, key_val))
        if keys:
            bulk_set_status(keys, export=False)

    return written


# ---------------------------------------------------------------------------
# Duplicate-tag prevention
#
# The Tag / Equip # is the identifier everything else keys off: the Index
# list, the PDF filename, and - critically - the Installed/Submitted status
# store (equipment_status.json is keyed by this value). Two rows sharing a
# Tag wouldn't just be a display annoyance; the status tracker would
# silently conflate two different physical pieces of equipment as one.
# ---------------------------------------------------------------------------
def find_duplicate_row(series_number, equip_key, key_value, exclude_row=None):
    """Returns the row number of another row already using this exact Tag /
    Equip # (case-insensitive), or None if it's free to use."""
    etype = EQUIPMENT_TYPES[equip_key]
    target = str(key_value).strip().casefold()
    if not target:
        return None
    for entry in read_index_rows(series_number, equip_key):
        if entry["row"] == exclude_row:
            continue
        existing = str(entry.get(etype["key_field"]) or "").strip().casefold()
        if existing == target:
            return entry["row"]
    return None


# ---------------------------------------------------------------------------
# Phase 13.5 - warn-don't-block validation. Unlike find_duplicate_row()
# above (a hard block - a shared Tag/Equip # would corrupt the status
# store), both of these are advisory: the value is saved regardless of
# what they return, they only tell the caller something's worth a second
# look.
# ---------------------------------------------------------------------------
TAG_SHAPE_RE = re.compile(r"^[A-Za-z0-9]+-[A-Za-z]+-[A-Za-z0-9]+$")


def tag_shape_warning(key_value):
    """None if key_value looks like AREA-TYPE-NUMBER (e.g.
    '29103-PIT-2171'); otherwise a short human-readable string. A blank
    value is out of scope here - that's find_duplicate_row/the required-
    field check's job, both of which DO block."""
    value = (key_value or "").strip()
    if not value or TAG_SHAPE_RE.match(value):
        return None
    return (f"“{value}” doesn't look like AREA-TYPE-NUMBER "
            f"(e.g. 29103-PIT-2171) - saved anyway.")


# Only the kind's single, primary identifying serial is checked - the
# common real case (one instrument, one serial). Transmitter has exactly
# one serial field; a valve's several component serials (actuator/
# positioner/solenoid/position_limit) are excluded from this check since
# those are frequently, legitimately reused parts across different valves.
SERIAL_FIELD_BY_KIND = {"transmitter": "serial_number", "valve": "valve_serial"}


def find_rows_with_duplicate_serial(equip_key, serial_value, exclude_series_row=None):
    """Every {'series', 'row', 'key_value'} across ALL registered series of
    this equipment kind whose primary serial field matches serial_value
    (case/space-insensitive) - a duplicate physical serial number is
    almost always a data-entry mistake (the same instrument entered
    twice, or two different ones sharing a typo'd serial), but never
    blocks a save the way a duplicate Tag does. exclude_series_row:
    (series_number, row_num) of the row being saved, so it never flags
    itself. A blank serial never counts as a duplicate of another blank
    one - reads every row's serial in ONE pass per series (reusing the
    header-derived column map), not a read_full_row() call per row."""
    field = SERIAL_FIELD_BY_KIND.get(equip_key)
    if not field:
        return []
    target = _normalize_for_compare(serial_value)
    if not target:
        return []

    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    key_field = etype["key_field"]
    matches = []
    for series_number in list_series():
        try:
            sheet_name = get_sheet_name(series_number, equip_key)
        except KeyError:
            continue
        wb = _get_cached_workbook(data_only=False)
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        field_to_col = export_mod.load_column_map(ws)
        serial_col = field_to_col.get(field)
        key_col = field_to_col.get(key_field)
        if not serial_col or not key_col:
            continue
        for r in range(export_mod.FIRST_DATA_ROW, ws.max_row + 1):
            if exclude_series_row == (series_number, r):
                continue
            raw_serial = export_mod.cell_to_str(ws.cell(row=r, column=serial_col).value)
            if _normalize_for_compare(raw_serial) != target:
                continue
            key_val = export_mod.cell_to_str(ws.cell(row=r, column=key_col).value)
            if key_val:
                matches.append({"series": series_number, "row": r, "key_value": key_val})
    return matches


# ---------------------------------------------------------------------------
# By-system breakdown, aggregated across every series - powers the Main
# Menu's dashboard (which otherwise only had one grand total per type).
# ---------------------------------------------------------------------------
def count_by_system_all_series(equip_key):
    """{'STEAM': 47, 'NATURAL GAS': 24, ..., '(No System Set)': 2} - the
    equipment type's System field, counted across every registered series
    combined. Rows with the key field filled in but no System value yet
    are bucketed under '(No System Set)' rather than silently dropped, so
    this always sums to exactly count_all_by_type()[equip_key] - the two
    numbers should never visibly disagree on screen."""
    etype = EQUIPMENT_TYPES[equip_key]
    system_field = etype["group_fields"][0]
    combined = {}
    unset = 0
    for series_number in list_series():
        try:
            rows = read_index_rows(series_number, equip_key)
        except KeyError:
            continue
        for entry in rows:
            value = str(entry.get(system_field) or "").strip()
            if value:
                combined[value] = combined.get(value, 0) + 1
            else:
                unset += 1
    result = dict(sorted(combined.items(), key=lambda kv: -kv[1]))  # busiest system first
    if unset:
        result["(No System Set)"] = unset
    return result


# ---------------------------------------------------------------------------
# Open the workbook itself directly (Settings -> "Open Excel File")
# ---------------------------------------------------------------------------
def open_workbook():
    open_file(WORKBOOK_PATH)


def open_sheet(series_number, equip_key):
    """Marks the given series+type's sheet as the workbook's active tab,
    saves that (a tiny metadata change - no data is touched), then opens
    the file with the OS's default app for .xlsx. Excel opens showing
    whichever sheet was active when the file was last saved, so this is
    a real jump-to-that-tab, not just 'open the workbook and hope'.

    Caveat worth knowing: if the workbook is ALREADY open in Excel, most
    Excel versions just bring the existing window to the front rather than
    re-reading the file from disk - so this can only land on a specific
    tab when nothing already has the file open."""
    sheet_name = get_sheet_name(series_number, equip_key)
    with _mutating_workbook() as wb:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Workbook has no sheet named '{sheet_name}'")
        wb.active = wb.sheetnames.index(sheet_name)
    open_file(WORKBOOK_PATH)


# ---------------------------------------------------------------------------
# Desktop shortcut creation. Windows-only (uses the WScript.Shell COM object
# via a one-line PowerShell call, which ships with every modern Windows -
# no extra Python dependency like pywin32 needed, which also keeps the
# PyInstaller build simpler).
# ---------------------------------------------------------------------------
def _shortcut_target_and_args():
    """(target_exe, arguments) the shortcut should launch."""
    if getattr(sys, "frozen", False):
        return sys.executable, ""
    # Running as a plain script: point at pythonw.exe (no console window)
    # with this script as its argument, falling back to python.exe if
    # pythonw isn't found next to the interpreter for some reason.
    pythonw = Path(sys.executable).with_name("pythonw.exe")
    target = str(pythonw) if pythonw.exists() else sys.executable
    script = str(HERE / "gui_app.py")
    return target, f'"{script}"'


def create_desktop_shortcut(shortcut_name="InstINDEX"):
    """Creates a .lnk on the Windows Desktop pointing at this app. Returns
    the shortcut path on success. Raises on any failure (no Desktop folder,
    not on Windows, PowerShell unavailable, etc.) - callers decide whether
    that's worth showing the user or failing silently."""
    if os.name != "nt":
        raise OSError("Desktop shortcuts are only supported on Windows.")

    desktop = Path(os.environ["USERPROFILE"]) / "Desktop"
    if not desktop.exists():
        raise FileNotFoundError(f"Couldn't find your Desktop folder at {desktop}")

    target, arguments = _shortcut_target_and_args()
    shortcut_path = desktop / f"{shortcut_name}.lnk"

    ps_script = (
        "$WshShell = New-Object -ComObject WScript.Shell;"
        f'$Shortcut = $WshShell.CreateShortcut("{shortcut_path}");'
        f'$Shortcut.TargetPath = "{target}";'
        f'$Shortcut.Arguments = \'{arguments}\';'
        f'$Shortcut.WorkingDirectory = "{HERE}";'
        f'$Shortcut.IconLocation = "{target}";'
        "$Shortcut.Save()"
    )
    subprocess.run(["powershell", "-NoProfile", "-NonInteractive", "-Command", ps_script],
                    check=True, capture_output=True, text=True)
    return shortcut_path


# ---------------------------------------------------------------------------
# Phase 14 - Global search (Ctrl+K). One flat, cached index across every
# registered series + both equipment kinds + the master list, so the
# palette dialog in gui_app.py never has to know how any of that data is
# actually stored - it just calls search_index(query).
# ---------------------------------------------------------------------------
SEARCH_INDEX_FIELDS = {
    "transmitter": ["tag", "system_number", "transmitter_type", "service",
                     "pid_number", "model", "serial_number"],
    "valve": ["equip_number", "system", "valve_type", "valve_model", "valve_serial",
              "actuator_serial", "positioner_serial", "solenoid_serial", "position_limit_serial"],
}


def _read_search_fields(series_number, equip_key):
    """Only the columns SEARCH_INDEX_FIELDS needs for this kind, read in
    ONE pass over the sheet (same batching reasoning as
    read_index_rows_with_progress - not a read_full_row() call per row,
    which would re-scan the header every time)."""
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    sheet_name = get_sheet_name(series_number, equip_key)
    key_field = etype["key_field"]

    wb = _get_cached_workbook(data_only=False)
    ws = wb[sheet_name]
    field_to_col = export_mod.load_column_map(ws)
    key_col = field_to_col.get(key_field)
    fields = SEARCH_INDEX_FIELDS.get(equip_key, [])

    rows = []
    for r in range(export_mod.FIRST_DATA_ROW, ws.max_row + 1):
        key_val = ws.cell(row=r, column=key_col).value if key_col else None
        if key_val in (None, ""):
            continue
        values = {fid: export_mod.cell_to_str(ws.cell(row=r, column=field_to_col.get(fid)).value)
                  if field_to_col.get(fid) else "" for fid in fields}
        values["row"] = r
        rows.append(values)
    return rows


def build_search_index():
    """[{'source': 'tracker'|'master', 'kind', 'series' (tracker only),
    'row' (tracker only), 'key_value', 'item' (master only), 'searchable',
    'label'}, ...] - a fresh, uncached build. Callers almost always want
    get_search_index() (cached) instead; this is the one-shot builder it
    calls, kept separate so it's independently testable."""
    entries = []
    for series_number in list_series():
        for equip_key, etype in EQUIPMENT_TYPES.items():
            try:
                rows = _read_search_fields(series_number, equip_key)
            except KeyError:
                continue
            for values in rows:
                key_value = values.get(etype["key_field"], "")
                searchable = " ".join(str(v) for v in values.values() if v).lower()
                entries.append({
                    "source": "tracker", "kind": equip_key, "series": series_number,
                    "row": values["row"], "key_value": key_value,
                    "searchable": searchable,
                    "label": f"{key_value} – {etype['label']} · "
                             f"{series_display_label(series_number)}",
                })

    for item in master_list_items():
        if item.get("kind") not in EQUIPMENT_TYPES:
            continue  # out_of_scope items aren't actionable from search (no create-from-master target)
        searchable = " ".join(str(item.get(k, "")) for k in
                               ("tag", "service", "model", "manufacturer", "type_desc")).lower()
        entries.append({
            "source": "master", "kind": item["kind"], "key_value": item["tag"], "item": item,
            "searchable": searchable,
            "label": f"{item['tag']} – {item.get('service') or item.get('type_desc') or ''} "
                     f"· Master list (not tracked)",
        })
    return entries


_SEARCH_INDEX_CACHE = {"mtime": None, "entries": None}


def get_search_index(force_refresh=False):
    """Cached across calls - rebuilt whenever the workbook's mtime has
    changed since the last build (covers every _mutating_workbook exit
    automatically, since that's exactly what changes the workbook's
    mtime) or when force_refresh is passed (a master-list re-import
    doesn't touch the workbook at all, so import_master_list() calls
    invalidate_search_index() explicitly instead)."""
    current_mtime = _workbook_mtime()
    if force_refresh or _SEARCH_INDEX_CACHE["entries"] is None or _SEARCH_INDEX_CACHE["mtime"] != current_mtime:
        _SEARCH_INDEX_CACHE["entries"] = build_search_index()
        _SEARCH_INDEX_CACHE["mtime"] = current_mtime
    return _SEARCH_INDEX_CACHE["entries"]


def invalidate_search_index():
    _SEARCH_INDEX_CACHE["entries"] = None
    _SEARCH_INDEX_CACHE["mtime"] = None


def search_index(query, limit=50):
    """The palette dialog's one entry point. [] for a blank query (never
    dumps the entire index) - substring match, case-insensitive, across
    tag/serials/model/service/system already flattened into each entry's
    'searchable' string."""
    query = (query or "").strip().lower()
    if not query:
        return []
    matches = [e for e in get_search_index() if query in e["searchable"]]
    return matches[:limit]


# ---------------------------------------------------------------------------
# Phase 16.1 - Automatic workbook backups.
#
# _backup_workbook_if_due() is called from _mutating_workbook()'s entry
# (above) - the one choke point every write already passes through.
# Never blocks a save: any failure here is warned to console and
# swallowed, same reasoning as write_json_atomic's sidecar writes.
# ---------------------------------------------------------------------------
BACKUPS_DIR = HERE / "backups"
REPORTS_DIR = HERE / "reports"
BACKUP_NAME_PREFIX = "Equipment_Inspection_Tracker."
BACKUP_NAME_SUFFIX = ".xlsx"


def _list_backup_paths():
    """Oldest first (by mtime) - [] if the folder doesn't exist yet."""
    if not BACKUPS_DIR.exists():
        return []
    paths = [p for p in BACKUPS_DIR.glob(f"{BACKUP_NAME_PREFIX}*{BACKUP_NAME_SUFFIX}") if p.is_file()]
    return sorted(paths, key=lambda p: p.stat().st_mtime)


def _write_backup_snapshot():
    """Copies the CURRENT on-disk workbook into backups/, timestamped to
    the second. Collision-safe (two backups in the same second get a
    '-2', '-3', ... suffix) so nothing is ever silently overwritten."""
    BACKUPS_DIR.mkdir(exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    dest = BACKUPS_DIR / f"{BACKUP_NAME_PREFIX}{stamp}{BACKUP_NAME_SUFFIX}"
    n = 2
    while dest.exists():
        dest = BACKUPS_DIR / f"{BACKUP_NAME_PREFIX}{stamp}-{n}{BACKUP_NAME_SUFFIX}"
        n += 1
    shutil.copy2(WORKBOOK_PATH, dest)
    return dest


def _prune_backups():
    keep = get_setting("backup_keep")
    if not keep or keep <= 0:
        return
    existing = _list_backup_paths()  # oldest first
    for old in existing[:-keep]:
        try:
            old.unlink()
        except OSError:
            pass


def _backup_workbook_if_due():
    if not WORKBOOK_PATH.exists():
        return  # nothing to back up yet (a brand-new install)
    try:
        interval_minutes = get_setting("backup_interval_minutes")
        if interval_minutes is None:
            interval_minutes = 30  # `or 30` would also override an explicit 0 ("always back up")
        existing = _list_backup_paths()
        if existing:
            newest = existing[-1]
            age_minutes = (datetime.datetime.now().timestamp() - newest.stat().st_mtime) / 60
            if age_minutes < interval_minutes:
                return
        _write_backup_snapshot()
        _prune_backups()
    except OSError as exc:
        print(f"WARNING: automatic workbook backup failed (continuing without backing up): {exc}")


def list_backups():
    """[{'path', 'name', 'mtime', 'size'}, ...], NEWEST first - the
    Backups dialog's list."""
    result = []
    for p in reversed(_list_backup_paths()):
        try:
            stat = p.stat()
        except OSError:
            continue
        result.append({"path": p, "name": p.name, "mtime": stat.st_mtime, "size": stat.st_size})
    return result


def backup_now():
    """The Backups dialog's 'Back up now' button - always writes a fresh
    snapshot regardless of backup_interval_minutes, then prunes to
    backup_keep same as the automatic path."""
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError("No workbook to back up yet.")
    dest = _write_backup_snapshot()
    _prune_backups()
    return dest


def restore_backup(backup_path):
    """Copies the chosen snapshot over the live workbook - AFTER taking
    ONE MORE safety snapshot of whatever's currently live, so restoring
    is itself undoable (by restoring that safety snapshot by hand).
    Returns the safety snapshot's path (None if there was no live
    workbook to protect - a restore onto a fresh install)."""
    backup_path = Path(backup_path)
    if not backup_path.exists():
        raise FileNotFoundError(f"Backup not found: {backup_path}")
    safety_snapshot = None
    if WORKBOOK_PATH.exists():
        safety_snapshot = _write_backup_snapshot()
    shutil.copy2(backup_path, WORKBOOK_PATH)
    invalidate_workbook_cache()
    invalidate_search_index()
    return safety_snapshot


# ---------------------------------------------------------------------------
# Phase 17 - Client progress report export.
#
# build_progress_report_data() is pure computation - zero new math. Every
# number in it comes straight from reconcile_master_list(),
# read_index_rows_with_progress(), and compute_progress()'s own milestone
# ids, the exact same functions Coverage and the Dashboard already call -
# so the report can never disagree with what's on screen (17's own
# acceptance criterion).
# ---------------------------------------------------------------------------
FLAG_LABELS = {
    "installed_mismatch": "Installed mismatch (master vs tracker)",
    "model_mismatch": "Model mismatch (master vs tracker)",
}


def find_all_duplicate_serials(equip_key):
    """Every group of 2+ rows (across ALL registered series) sharing the
    same primary serial value for this kind - the whole-tracker version
    of find_rows_with_duplicate_serial()'s single-value check, used by
    the report's Flags sheet. [{'serial', 'rows': [{'series','row',
    'key_value'}, ...]}, ...]. One pass per series/sheet, same as
    find_rows_with_duplicate_serial."""
    field = SERIAL_FIELD_BY_KIND.get(equip_key)
    if not field:
        return []
    etype = EQUIPMENT_TYPES[equip_key]
    export_mod = etype["export_module"]
    key_field = etype["key_field"]

    by_serial = {}
    for series_number in list_series():
        try:
            sheet_name = get_sheet_name(series_number, equip_key)
        except KeyError:
            continue
        wb = _get_cached_workbook(data_only=False)
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        field_to_col = export_mod.load_column_map(ws)
        serial_col = field_to_col.get(field)
        key_col = field_to_col.get(key_field)
        if not serial_col or not key_col:
            continue
        for r in range(export_mod.FIRST_DATA_ROW, ws.max_row + 1):
            raw_serial = export_mod.cell_to_str(ws.cell(row=r, column=serial_col).value)
            target = _normalize_for_compare(raw_serial)
            if not target:
                continue
            key_val = export_mod.cell_to_str(ws.cell(row=r, column=key_col).value)
            if not key_val:
                continue
            by_serial.setdefault(target, {"serial": raw_serial.strip(), "rows": []})
            by_serial[target]["rows"].append({"series": series_number, "row": r, "key_value": key_val})

    return [group for group in by_serial.values() if len(group["rows"]) > 1]


def build_progress_report_data():
    """Returns {'has_master_list': bool, 'summary': [...], 'missing':
    [...], 'flags': [...]}. summary rows: one per area x kind when a
    master list has been imported (in-scope = matched + missing, exactly
    Coverage's own definition); one per series x kind (tracker-only,
    noted as such) when it hasn't - 17's own 'degrades to tracker-only
    summary with a note' requirement."""
    snapshot = load_master_list()
    reconciliation = reconcile_master_list() if snapshot else None

    summary = []
    missing = []
    flags = []

    if reconciliation:
        groups = {}
        for pair in reconciliation["matched"]:
            item = pair["master"]
            groups.setdefault((item["area"], item["kind"]), {"matched": [], "missing": []})
            groups[(item["area"], item["kind"])]["matched"].append(pair)
        for item in reconciliation["missing"]:
            groups.setdefault((item["area"], item["kind"]), {"matched": [], "missing": []})
            groups[(item["area"], item["kind"])]["missing"].append(item)

        for (area, kind), bucket in sorted(groups.items()):
            matched, miss = bucket["matched"], bucket["missing"]
            total = len(matched) + len(miss)
            milestone_counts = {mid: 0 for mid, _label in PROGRESS_MILESTONES}
            percents = []
            for pair in matched:
                percents.append(pair["progress"]["percent"])
                for m in pair["progress"]["milestones"]:
                    if m["done"]:
                        milestone_counts[m["id"]] += 1
            summary.append({
                "area": area, "kind": kind, "total_in_scope": total,
                "created": milestone_counts["created"],
                "serials_captured": milestone_counts["serial_captured"],
                "installed": milestone_counts["installed"],
                "inspection_complete": milestone_counts["inspection_complete"],
                "submitted": milestone_counts["submitted"],
                "accepted": milestone_counts["accepted"],
                "avg_percent": round(sum(percents) / len(percents)) if percents else 0,
            })

        for item in reconciliation["missing"]:
            missing.append({
                "tag": item["tag"], "service": item.get("service", ""),
                "type_desc": item.get("type_desc", ""), "sheet": item.get("source_sheet", ""),
            })

        for pair in reconciliation["matched"]:
            if pair["flags"]:
                flags.append({
                    "tag": pair["master"]["tag"], "kind": pair["master"]["kind"],
                    "detail": "; ".join(FLAG_LABELS.get(f, f) for f in pair["flags"]),
                })
    else:
        for series_number in list_series():
            for equip_key, etype in EQUIPMENT_TYPES.items():
                try:
                    rows = read_index_rows_with_progress(series_number, equip_key)
                except KeyError:
                    continue
                if not rows:
                    continue
                milestone_counts = {mid: 0 for mid, _label in PROGRESS_MILESTONES}
                percents = []
                for entry in rows:
                    percents.append(entry["progress"]["percent"])
                    for m in entry["progress"]["milestones"]:
                        if m["done"]:
                            milestone_counts[m["id"]] += 1
                summary.append({
                    "area": f"{series_display_label(series_number)} (tracker only)", "kind": equip_key,
                    "total_in_scope": len(rows),
                    "created": milestone_counts["created"],
                    "serials_captured": milestone_counts["serial_captured"],
                    "installed": milestone_counts["installed"],
                    "inspection_complete": milestone_counts["inspection_complete"],
                    "submitted": milestone_counts["submitted"],
                    "accepted": milestone_counts["accepted"],
                    "avg_percent": round(sum(percents) / len(percents)) if percents else 0,
                })

    # 13.5 validation warnings, across the whole tracker - independent of
    # whether a master list has been imported.
    for equip_key, etype in EQUIPMENT_TYPES.items():
        key_field = etype["key_field"]
        for series_number in list_series():
            try:
                rows = read_index_rows(series_number, equip_key)
            except KeyError:
                continue
            for r in rows:
                key_val = r.get(key_field, "")
                warning = tag_shape_warning(key_val)
                if warning:
                    flags.append({"tag": key_val, "kind": equip_key, "detail": warning})
        for group in find_all_duplicate_serials(equip_key):
            tags = ", ".join(f"{row['key_value']} (series {row['series']})" for row in group["rows"])
            flags.append({
                "tag": tags, "kind": equip_key,
                "detail": f"Duplicate serial \"{group['serial']}\"",
            })

    return {"has_master_list": bool(snapshot), "summary": summary, "missing": missing, "flags": flags}


def write_progress_report_xlsx(data, out_path):
    """Renders build_progress_report_data()'s dict to a standalone
    workbook - openpyxl only, no new dependency. Modest formatting: bold
    header row, autofilter, frozen header row, and the workbook's own
    Good/Bad palette (reused from row_status colors' hex values, kept in
    sync by hand since theme.py's Qt colors and this file's openpyxl
    fills are two different color systems with no shared source)."""
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
    BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
    HEADER_FONT = Font(bold=True)

    wb = openpyxl.Workbook()

    def _write_sheet(ws, headers, rows, pct_col=None, header_row=1):
        """header_row lets Sheet 1 carry a note above the header (the
        tracker-only degrade case) without duplicating this whole
        function - everything below just shifts down."""
        for c, text in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=c, value=text)
            cell.font = HEADER_FONT
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
        for r, row in enumerate(rows, start=header_row + 1):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
        if pct_col is not None:
            for r in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(row=r, column=pct_col)
                cell.fill = GOOD_FILL if isinstance(cell.value, (int, float)) and cell.value >= 100 else BAD_FILL
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 22

    ws1 = wb.active
    ws1.title = "Summary"
    header_row = 1
    if not data["has_master_list"]:
        ws1.append(["No master list has been imported - this report is tracker-only "
                     "(every row's own scope, not reconciled against the client's master list)."])
        ws1.append([])
        header_row = 3
    headers1 = ["Area", "Kind", "Total In Scope", "Created", "Serials Captured",
                "Installed", "Inspection Complete", "Submitted", "Accepted", "Avg %"]
    rows1 = [[r["area"], r["kind"], r["total_in_scope"], r["created"], r["serials_captured"],
              r["installed"], r["inspection_complete"], r["submitted"], r["accepted"], r["avg_percent"]]
             for r in data["summary"]]
    _write_sheet(ws1, headers1, rows1, pct_col=len(headers1), header_row=header_row)

    ws2 = wb.create_sheet("Missing")
    _write_sheet(ws2, ["Tag", "Service", "Type Description", "Sheet"],
                 [[r["tag"], r["service"], r["type_desc"], r["sheet"]] for r in data["missing"]])

    ws3 = wb.create_sheet("Flags")
    _write_sheet(ws3, ["Tag(s)", "Kind", "Detail"],
                 [[r["tag"], r["kind"], r["detail"]] for r in data["flags"]])

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def export_progress_report():
    """Builds the report and saves it to reports/Progress_Report_<date>.
    xlsx next to the workbook. Returns the path written."""
    data = build_progress_report_data()
    stamp = datetime.date.today().isoformat()
    out_path = REPORTS_DIR / f"Progress_Report_{stamp}.xlsx"
    n = 2
    while out_path.exists():
        out_path = REPORTS_DIR / f"Progress_Report_{stamp}-{n}.xlsx"
        n += 1
    return write_progress_report_xlsx(data, out_path)


# ---------------------------------------------------------------------------
# Phase 18 - "Cleaned copy" export.
#
# Never opens Equipment_Inspection_Tracker.xlsx itself for writing - reads
# it through the normal cached read path (same as every report/validation
# function above) to work out what a cleanup WOULD do, then makes a
# byte-identical copy (shutil.copy2, not an openpyxl load/re-save - that
# keeps every data-validation dropdown, conditional-formatting rule, and
# style exactly as they are) and only opens THAT copy to apply cell-value
# fixes. Never inserts, deletes, or reorders rows - same reason the
# "Removing rows" comment above gives for clearing a row in place instead
# of ws.delete_rows(): data-validation ranges and conditional-formatting
# rules are anchored to specific row numbers, and openpyxl's support for
# re-anchoring those on a shift is known to be incomplete. A truly blank
# row is left exactly as-is - indistinguishable from unused capacity, same
# as save_row leaves a cleared one.
#
# Two kinds of cleanup, and the line between them is never guessed:
#   - Fixed automatically (always safe, always deterministic): leading/
#     trailing whitespace trimmed off every non-date, string-typed schema
#     field, plus the row's key field (Tag / Equip #) upper-cased too. A
#     cell is only written if the raw value actually changes, and only
#     when it's already a plain str - a numeric/bool/date-typed cell is
#     never touched, so its type and number format always survive. This
#     also means the comparison is against the RAW cell value, not
#     cell_to_str()'s output: that helper already strips whitespace itself
#     (str(value).strip()) and reformats other types (a float cell's
#     "2171.0" comes back as "2171"), so diffing against it would both
#     hide real whitespace and misreport an ordinary numeric-type
#     difference as a whitespace "fix". A formula cell's raw value is
#     ALSO a plain str with data_only=False (its formula text, not its
#     computed result) - excluded too, so a string literal embedded
#     inside a formula is never silently rewritten.
#   - Flagged, never touched: a duplicate primary serial
#     (find_all_duplicate_serials) or a tag that still doesn't look like
#     AREA-TYPE-NUMBER after trimming/upper-casing (tag_shape_warning) -
#     both already-trusted checks the Progress Report's Flags sheet uses,
#     reused here rather than reinvented, so the two can't disagree with
#     each other about what counts as an issue.
# Both are also listed on a new "Cleanup Log" sheet appended to the copy,
# and lightly highlighted in place on the copy's own data sheets, so
# either is visible without leaving the workbook.
# ---------------------------------------------------------------------------
CLEANED_DIR = HERE / "cleaned"


def build_cleanup_plan():
    """Dry-run over the live workbook (read-only, never opens it for
    writing) - {'fixes': [...], 'flags': [...]}. Safe to call any time,
    including just to preview what a cleanup would find."""
    fixes = []
    flags = []
    for equip_key, etype in EQUIPMENT_TYPES.items():
        export_mod = etype["export_module"]
        key_field = etype["key_field"]
        date_fields = set(etype.get("date_fields", []))
        for series_number in list_series():
            try:
                sheet_name = get_sheet_name(series_number, equip_key)
            except KeyError:
                continue
            wb = _get_cached_workbook(data_only=False)
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            field_to_col = export_mod.load_column_map(ws)
            key_col = field_to_col.get(key_field)
            if not key_col:
                continue
            for r in range(export_mod.FIRST_DATA_ROW, ws.max_row + 1):
                # "Is this row used" uses the exact same raw (None, "")
                # test as read_index_rows()/find_first_blank_row() - NOT
                # cell_to_str()'s stripped output, which would collapse a
                # whitespace-only tag (a real, visible row everywhere else
                # in this app) to "" and wrongly skip the whole row here,
                # including every one of its OTHER fields that might
                # genuinely need a whitespace fix.
                raw_key_val = ws.cell(row=r, column=key_col).value
                if raw_key_val in (None, ""):
                    continue  # unused row - left exactly as-is
                key_val = export_mod.cell_to_str(raw_key_val)
                for fid, col in field_to_col.items():
                    if fid in date_fields:
                        continue  # never rewrite a date cell's type/format
                    raw = ws.cell(row=r, column=col).value
                    # Only ever touch a cell that's actually a Python str -
                    # cell_to_str() itself already strips whitespace before
                    # a caller sees it (str(value).strip()), so comparing
                    # against ITS output would silently mask real
                    # whitespace and, worse, would flag a plain type
                    # difference (e.g. a float cell's "2171.0" vs.
                    # cell_to_str's "2171") as a "fix" - a type change, not
                    # a whitespace trim. Reading the raw value directly and
                    # skipping anything that isn't already a str sidesteps
                    # both: a numeric/bool/date-typed cell is never
                    # rewritten by this path at all. A formula cell's raw
                    # value is ALSO a plain str (its formula text, e.g.
                    # "=CONCATENATE(...)") with data_only=False, so it's
                    # excluded explicitly too - trimming/upper-casing that
                    # text would silently rewrite a string literal INSIDE
                    # the formula, changing what it actually computes.
                    if not isinstance(raw, str) or raw.startswith("="):
                        continue
                    old_text = raw
                    new_text = raw.strip()
                    if fid == key_field:
                        new_text = new_text.upper()
                    if new_text != old_text:
                        fixes.append({
                            "series": series_number, "equip_key": equip_key, "row": r,
                            "sheet": sheet_name, "col": col, "field": fid,
                            "key_value": key_val.strip().upper(),
                            "old": old_text, "new": new_text,
                        })
                final_key = key_val.strip().upper()
                warning = tag_shape_warning(final_key)
                if warning:
                    flags.append({
                        "series": series_number, "equip_key": equip_key, "row": r,
                        "sheet": sheet_name, "key_value": final_key, "detail": warning,
                    })

        for group in find_all_duplicate_serials(equip_key):
            # find_all_duplicate_serials()'s own key_value is only
            # whitespace-stripped (via cell_to_str), not upper-cased - it's
            # shared with the Progress Report's Flags sheet (Phase 17),
            # which has its own tests expecting that exact behavior, so
            # it's never changed here. Normalized separately, just for
            # this flag's own display, so it matches the casing every
            # OTHER Cleanup Log entry (and the actual fixed cell, if this
            # same row's key also got trimmed/upper-cased above) uses.
            tag_list = ", ".join(
                f"{row['key_value'].strip().upper()} (series {row['series']})" for row in group["rows"]
            )
            for row in group["rows"]:
                flags.append({
                    "series": row["series"], "equip_key": equip_key, "row": row["row"],
                    "sheet": get_sheet_name(row["series"], equip_key),
                    "key_value": row["key_value"].strip().upper(),
                    "detail": f"Duplicate serial \"{group['serial']}\" - shared with {tag_list}",
                })

    return {"fixes": fixes, "flags": flags}


def _apply_cleanup_plan(out_wb, plan):
    """Mutates out_wb (a freshly-loaded copy, never the live workbook) in
    place: writes each fix's new value, then highlights every touched
    cell - amber for a plain fix, red for a flagged-only key cell, and a
    third color for a key cell that's BOTH (trimmed/upper-cased by a fix
    AND still flagged afterward - e.g. a bad tag shape that whitespace
    alone can't cure) so that overlap is never silently lost to whichever
    color happened to be applied second. Cell-value and cell-fill writes
    only - no row insert/delete, so nothing anchored to a row number
    (data validation, conditional formatting) is ever disturbed."""
    from openpyxl.styles import PatternFill
    FIXED_FILL = PatternFill("solid", fgColor="FFEB9C")       # light amber - "changed by cleanup"
    FLAG_FILL = PatternFill("solid", fgColor="FFC7CE")        # same red as the Progress Report's BAD_FILL
    FIXED_AND_FLAG_FILL = PatternFill("solid", fgColor="FFD966")  # amber+red blend - both apply

    fixed_by_cell = {(fix["sheet"], fix["row"], fix["col"]): fix for fix in plan["fixes"]}

    flagged_key_cell_by_row = {}
    for flag in plan["flags"]:
        row_key = (flag["sheet"], flag["row"])
        if row_key in flagged_key_cell_by_row:
            continue
        etype = EQUIPMENT_TYPES[flag["equip_key"]]
        export_mod = etype["export_module"]
        ws = out_wb[flag["sheet"]]
        key_col = export_mod.load_column_map(ws).get(etype["key_field"])
        if key_col:
            flagged_key_cell_by_row[row_key] = key_col
    flagged_cells = {(sheet, row, col) for (sheet, row), col in flagged_key_cell_by_row.items()}

    for cell_id, fix in fixed_by_cell.items():
        sheet, row, col = cell_id
        cell = out_wb[sheet].cell(row=row, column=col)
        cell.value = fix["new"]
        cell.fill = FIXED_AND_FLAG_FILL if cell_id in flagged_cells else FIXED_FILL

    for cell_id in flagged_cells:
        if cell_id in fixed_by_cell:
            continue  # already colored FIXED_AND_FLAG_FILL above
        sheet, row, col = cell_id
        out_wb[sheet].cell(row=row, column=col).fill = FLAG_FILL


def _write_cleanup_log_sheet(wb, plan):
    """Appends a new sheet named 'Cleanup Log' (or, on the rare chance the
    live workbook already has one - e.g. a previous cleaned copy was
    promoted to become the new live tracker - a collision-suffixed name
    like a repeat backup gets: never silently reused, renamed, or
    overwritten, so an old log is never mistaken for this run's) after
    every existing sheet in wb - original sheet order/tabs are otherwise
    untouched."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    HEADER_FONT = Font(bold=True)

    title = "Cleanup Log"
    n = 2
    while title in wb.sheetnames:
        title = f"Cleanup Log ({n})"
        n += 1
    ws = wb.create_sheet(title)
    r = 1
    ws.cell(row=r, column=1, value="Auto-fixed cells").font = HEADER_FONT
    r += 1
    for c, text in enumerate(
        ["Series", "Sheet", "Row", "Tag/Equip #", "Field", "Old value", "New value"], start=1
    ):
        ws.cell(row=r, column=c, value=text).font = HEADER_FONT
    r += 1
    if not plan["fixes"]:
        ws.cell(row=r, column=1, value="(nothing to fix - already clean)")
        r += 1
    for fix in plan["fixes"]:
        for c, value in enumerate(
            [fix["series"], fix["sheet"], fix["row"], fix["key_value"], fix["field"], fix["old"], fix["new"]],
            start=1,
        ):
            ws.cell(row=r, column=c, value=value)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Needs manual review (not auto-fixed)").font = HEADER_FONT
    r += 1
    for c, text in enumerate(["Series", "Sheet", "Row", "Tag/Equip #", "Issue"], start=1):
        ws.cell(row=r, column=c, value=text).font = HEADER_FONT
    r += 1
    if not plan["flags"]:
        ws.cell(row=r, column=1, value="(nothing flagged - already clean)")
        r += 1
    for flag in plan["flags"]:
        for c, value in enumerate(
            [flag["series"], flag["sheet"], flag["row"], flag["key_value"], flag["detail"]], start=1
        ):
            ws.cell(row=r, column=c, value=value)
        r += 1

    for c in range(1, 8):
        ws.column_dimensions[get_column_letter(c)].width = 26


def export_cleaned_workbook():
    """Writes cleaned/Equipment_Inspection_Tracker_CLEANED_<date>.xlsx next
    to the live workbook - collision-suffixed like export_progress_report(),
    so a same-day re-export never clobbers an earlier one. The live
    workbook (WORKBOOK_PATH) is only ever read by this function, never
    opened for writing. Returns the path written."""
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError("No workbook to clean yet.")
    plan = build_cleanup_plan()

    CLEANED_DIR.mkdir(exist_ok=True)
    stamp = datetime.date.today().isoformat()
    out_path = CLEANED_DIR / f"Equipment_Inspection_Tracker_CLEANED_{stamp}.xlsx"
    n = 2
    while out_path.exists():
        out_path = CLEANED_DIR / f"Equipment_Inspection_Tracker_CLEANED_{stamp}-{n}.xlsx"
        n += 1
    shutil.copy2(WORKBOOK_PATH, out_path)

    out_wb = openpyxl.load_workbook(out_path)
    _apply_cleanup_plan(out_wb, plan)
    _write_cleanup_log_sheet(out_wb, plan)
    out_wb.save(out_path)
    return out_path
